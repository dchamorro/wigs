"""
test_sync.py — Verifica que sync_from_airtable.py vuelca las entradas del bundle
en las celdas correctas (las mismas que lee el parser / valida test_contract).

Construye un tablero en blanco, aplica el fixture y comprueba que dueño, meta,
nombres/metas de leads, lecturas semanales (emparejadas por fecha), backlog por
año y el bloque NAT del Dashboard cayeron donde deben.

Uso:  python3 tests/test_sync.py   (requiere openpyxl; no necesita recalc)
"""
import os
import sys
import tempfile
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'excel'))

from openpyxl import load_workbook
from migrate_data import date_map, _find_annual_nat, _find_key_rows  # noqa: E402
import sync_from_airtable as sync  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FIXTURE = os.path.join(ROOT, 'excel', 'airtable_sample.json')


def _build_blank(path):
    """build_wig.py guarda en sys.argv[1] al ejecutarse como script; aquí lo
    invocamos importándolo con argv preparado."""
    import runpy
    sys.argv = ['build_wig.py', path]
    runpy.run_path(os.path.join(ROOT, 'excel', 'build_wig.py'), run_name='__main__')


def main():
    with tempfile.TemporaryDirectory() as td:
        blank = os.path.join(td, 'blank.xlsx')
        out = os.path.join(td, 'out.xlsx')
        _build_blank(blank)
        bundle = sync.from_fixture(FIXTURE)
        sync.apply_to_workbook(bundle, blank, out)

        wb = load_workbook(out, data_only=False)

        # WIG semanal: cabecera + leads + lecturas por fecha
        ws = wb['1. Smart User 500']
        assert ws['B4'].value == 'María José Selva', ws['B4'].value
        assert ws['B5'].value == 500, ws['B5'].value
        assert ws.cell(row=8, column=9).value == 'L1. Cartas de asignación firmadas'
        assert ws.cell(row=9, column=9).value == 20
        nmap = date_map(ws)
        r = nmap[date(2026, 6, 22)]
        assert ws.cell(row=r, column=4).value == 48, 'lag real 2026-06-22'
        assert ws.cell(row=r, column=9).value == 14, 'L1 real 2026-06-22'
        assert ws.cell(row=r, column=11).value == 2, 'L2 real 2026-06-22'

        # WIG mensual
        wm = wb['8. GP 16%']
        assert wm['B5'].value == 0.16
        rm = date_map(wm)[date(2026, 7, 1)]
        assert abs(wm.cell(row=rm, column=4).value - 0.142) < 1e-9, 'lag mensual'

        # Página por año: backlog (col B) por fecha + GP% supuesto (2029)
        y = wb['2027']
        amap, rr = {}, 12
        while y.cell(row=rr, column=1).value is not None:
            v = y.cell(row=rr, column=1).value
            amap[v.date() if hasattr(v, 'date') else v] = rr
            rr += 1
        assert y.cell(row=amap[date(2026, 6, 22)], column=2).value == 1280000
        assert wb['2029']['E4'].value == 0.17, 'GP% supuesto 2029'

        # Dashboard: meta anual NAT, NAT real mensual, NAT real por año
        d = wb['Dashboard']
        ar = _find_annual_nat(d)
        assert d.cell(row=ar[0], column=ar[1]).value == 1000000, 'meta anual NAT'
        meses = _find_key_rows(d, 'Mes')
        assert d.cell(row=meses[date(2027, 1, 1)], column=3).value == 70000, 'NAT mensual ene-27'
        # trayectoria '2027 (meta)' por los 4 dígitos
        yrow = next(row for lbl, row in _find_key_rows(d, 'Año').items() if '2027' in str(lbl))
        assert d.cell(row=yrow, column=6).value == 90000, 'NAT real 2027'

    print('OK  test_sync — todas las entradas cayeron en las celdas del contrato')


if __name__ == '__main__':
    main()
