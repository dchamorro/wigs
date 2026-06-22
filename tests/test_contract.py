"""
test_contract.py — Valida el CONTRATO entre el Excel y el parser del marcador.

El marcador (web/marcador.html) lee posiciones FIJAS del workbook. Si este
test falla después de un cambio en excel/build_wig.py, el televisor se
rompería: hay que ajustar también el parser en marcador.html (función
parseWorkbook) y mantener ambos lados sincronizados.

Uso:  python3 -m pytest tests/ -q        (o)        python3 tests/test_contract.py
Requiere un tablero construido y recalculado en dist/Tablero_WIG_4DX_GBM.xlsx
"""
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook

WB_PATH = os.environ.get('WIG_XLSX', 'dist/Tablero_WIG_4DX_GBM.xlsx')
HTML_PATH = os.environ.get('WIG_HTML', 'web/marcador.html')
DATA0 = 11
MAX_LEADS = 8


def wb():
    assert os.path.exists(WB_PATH), f'Construir primero: make build (falta {WB_PATH})'
    return load_workbook(WB_PATH, data_only=False)


NO_WIG = ('Dashboard', 'Instrucciones', 'Compromisos', 'Tareas')


def is_year_page(name):
    """Pestaña por año: '2027'/'2028'/'2029' (o el legado 'Backlog <año>')."""
    return bool(re.match(r'^\d{4}$', name)) or name.startswith('Backlog ')


def is_support(name):
    """No es una pestaña de WIG (Dashboard/Instrucciones/Compromisos/Tareas o página por año)."""
    return name in NO_WIG or is_year_page(name)


def test_tabs_exist():
    names = wb().sheetnames
    assert names[0] == 'Dashboard'
    assert 'Instrucciones' in names
    wig_tabs = [n for n in names if not is_support(n)]
    assert len(wig_tabs) >= 1, 'Debe existir al menos una pestaña de WIG'


def test_wig_tab_layout():
    book = wb()
    for name in book.sheetnames:
        if is_support(name):
            continue
        ws = book[name]
        # encabezado que el parser lee
        assert ws['A1'].value, f'{name}: falta título en A1'
        assert ws['B2'].value, f'{name}: falta lag en B2'
        assert ws['B4'].value is not None, f'{name}: falta dueño en B4'
        assert ws['B5'].value is not None, f'{name}: falta meta en B5'
        # detección acum/nivel: E10 con o sin texto
        e10 = ws['E10'].value
        acum = e10 is not None and str(e10).strip() != ''
        # primera fila de datos: fecha en B11 y fórmulas correctas
        assert ws.cell(row=DATA0, column=2).value is not None, f'{name}: falta fecha en B{DATA0}'
        estado = ws.cell(row=DATA0, column=8).value
        assert isinstance(estado, str) and estado.startswith('='), f'{name}: H{DATA0} debe ser fórmula de Estado'
        if acum:
            for c, frag in ((5, 'SUM'), (6, 'SUM'), (7, '/')):
                v = ws.cell(row=DATA0, column=c).value
                assert isinstance(v, str) and v.startswith('=') and frag in v, \
                    f'{name}: col {c} fila {DATA0} no tiene la fórmula esperada'
        # bloque de leads: nombre fila 8, meta fila 9, % en columna par
        found = 0
        for k in range(MAX_LEADS):
            c1 = 9 + 2 * k
            nm = ws.cell(row=8, column=c1).value
            if nm and '(Disponible)' not in str(nm):
                found += 1
                pct = ws.cell(row=DATA0, column=c1 + 1).value
                assert isinstance(pct, str) and pct.startswith('='), \
                    f'{name}: L{k+1} sin fórmula de % en fila {DATA0}'
        assert found >= 1, f'{name}: ningún lead activo en fila 8'


def test_compromisos_layout():
    """Compromisos es OPCIONAL (tableros previos no la tienen y el parser la
    trata como opcional). Si existe, sus encabezados deben ser los del contrato."""
    book = wb()
    if 'Compromisos' not in book.sheetnames:
        return
    ws = book['Compromisos']
    hdrs = [ws.cell(row=1, column=c).value for c in range(1, 7)]
    assert hdrs == ['Semana', 'WIG', 'Lead', 'Compromiso', 'Responsable', 'Estado'], \
        f'Compromisos: encabezados inesperados {hdrs} — el parser del marcador los lee por posición A–F'


def test_tareas_layout():
    """Tareas (tareas de soporte por lead): tabla plana A:E con encabezados fijos
    en fila 1; el parser del marcador la lee por posición A–E (WIG, Lead, Tarea,
    Responsable, Estado). El build la genera siempre."""
    book = wb()
    assert 'Tareas' in book.sheetnames, 'falta la pestaña Tareas'
    ws = book['Tareas']
    hdrs = [ws.cell(row=1, column=c).value for c in range(1, 6)]
    assert hdrs == ['WIG', 'Lead', 'Tarea', 'Responsable', 'Estado'], \
        f'Tareas: encabezados inesperados {hdrs} — el parser los lee por posición A–E'
    # la semilla deja al menos una fila de ejemplo con WIG/Lead/Tarea
    assert ws.cell(row=2, column=1).value is not None, 'Tareas: falta WIG en A2'
    assert ws.cell(row=2, column=2).value is not None, 'Tareas: falta Lead en B2'
    assert ws.cell(row=2, column=3).value, 'Tareas: falta Tarea en C2'


def test_dashboard_layout():
    ws = wb()['Dashboard']
    # Trayectoria NAT: 4 años desde fila 6 (2026–2029)
    assert ws.cell(row=6, column=1).value and ws.cell(row=9, column=1).value, \
        'Dashboard: la trayectoria debe tener 4 años (filas 6–9)'
    # Cobertura de backlog: encabezado en fila 13, datos 14–16
    assert ws.cell(row=13, column=1).value == 'Año', 'Dashboard: falta tabla de backlog (fila 13)'
    f_bl = ws.cell(row=14, column=2).value
    assert isinstance(f_bl, str) and f_bl.startswith('='), 'Dashboard: B14 debe referenciar la página Backlog'
    # Seguimiento mensual NAT: meta anual en C19, meses desde fila 22, acum en col D
    assert ws['C19'].value is not None, 'Dashboard: falta meta anual en C19'
    assert ws.cell(row=22, column=1).value is not None, 'Dashboard: falta primer mes en A22'
    f = ws.cell(row=22, column=4).value
    assert isinstance(f, str) and f.startswith('='), 'Dashboard: D22 debe ser fórmula'


def test_year_pages_layout():
    """Páginas por año (2027/2028/2029): meta NAT (E5), meta de GP (B5 = Ingreso ×
    GP%) y tabla semanal de GP comprometido (col B, input) — el Dashboard las
    referencia por posición (B5/B6/B7/E6)."""
    book = wb()
    yp = [n for n in book.sheetnames if is_year_page(n)]
    assert len(yp) >= 1, 'Debe existir al menos una página por año'
    for name in yp:
        ws = book[name]
        assert ws['E5'].value is not None, f'{name}: falta la meta NAT en E5'
        b5 = ws['B5'].value
        assert isinstance(b5, str) and b5.startswith('='), f'{name}: B5 (GP meta) debe ser fórmula'
        assert ws.cell(row=12, column=1).value is not None, f'{name}: falta primer semana en A12'
        # col C de la tabla = referencia a la meta de GP
        c = ws.cell(row=12, column=3).value
        assert isinstance(c, str) and c.startswith('='), f'{name}: C12 (GP meta por fila) debe ser fórmula'


def test_html_parser_matches():
    """El parser del HTML debe leer las mismas celdas ancla del contrato."""
    assert os.path.exists(HTML_PATH), f'falta {HTML_PATH}'
    html = open(HTML_PATH, encoding='utf-8').read()
    for anchor in ("g('B2')", "g('B4')", "g('B5')", "g('E10')", "g('B'+r)",
                   "g('C19')", "encode_col(8+2*k)", "'Compromisos'", "'Tareas'", "parseTareas"):
        assert anchor in html, f'parser: ancla {anchor} no encontrada — contrato roto'
    assert "const CONFIG = { DATA_URL: ''" in html, (
        'falta la línea CONFIG exacta — el workflow de Azure parchea esa cadena literal')


def test_recalculated():
    """SheetJS lee valores cacheados: el workbook publicado debe estar recalculado."""
    book = load_workbook(WB_PATH, data_only=True)
    name = [n for n in book.sheetnames if not is_support(n)][0]
    ws = book[name]
    v = ws.cell(row=DATA0, column=3).value  # Meta: fórmula en tabs acum, constante en nivel
    assert v is not None, ('Sin valores cacheados: ejecutar python3 scripts/recalc.py ' + WB_PATH)


if __name__ == '__main__':
    fns = [v for k, v in sorted(globals().items()) if k.startswith('test_')]
    for fn in fns:
        fn()
        print(f'OK  {fn.__name__}')
    print(f'\n{len(fns)} pruebas de contrato pasaron.')
