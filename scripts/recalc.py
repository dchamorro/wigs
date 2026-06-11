"""
recalc.py — Recalcula las fórmulas de un .xlsx con LibreOffice headless.

¿Por qué es obligatorio? openpyxl escribe las fórmulas pero NO sus valores
calculados, y el marcador (SheetJS) lee los valores cacheados. Un tablero
construido o migrado sin recalcular se ve VACÍO en el televisor.

Uso:  python3 scripts/recalc.py archivo.xlsx
Requiere: LibreOffice (sudo apt install libreoffice-calc / brew install --cask libreoffice)
"""
import os
import platform
import shutil
import subprocess
import sys
from pathlib import Path

MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

MACRO_DIR = (
    "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
    if platform.system() == "Darwin"
    else "~/.config/libreoffice/4/user/basic/Standard"
)


def soffice():
    for cand in ("soffice", "/Applications/LibreOffice.app/Contents/MacOS/soffice"):
        if shutil.which(cand) or os.path.exists(cand):
            return cand
    sys.exit("LibreOffice no encontrado. Instalar libreoffice-calc.")


def ensure_macro(bin_):
    d = os.path.expanduser(MACRO_DIR)
    f = os.path.join(d, "Module1.xba")
    if os.path.exists(f) and "RecalculateAndSave" in Path(f).read_text():
        return
    if not os.path.isdir(d):
        subprocess.run([bin_, "--headless", "--terminate_after_init"],
                       capture_output=True, timeout=30)
        os.makedirs(d, exist_ok=True)
    Path(f).write_text(MACRO)
    script_xlb = os.path.join(d, "script.xlb")
    if not os.path.exists(script_xlb):
        Path(script_xlb).write_text(
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">\n'
            '<library:library xmlns:library="http://openoffice.org/2000/library" library:name="Standard" '
            'library:readonly="false" library:passwordprotected="false">\n'
            ' <library:element library:name="Module1"/>\n</library:library>')


def main(path):
    path = os.path.abspath(path)
    if not os.path.exists(path):
        sys.exit(f"No existe: {path}")
    bin_ = soffice()
    ensure_macro(bin_)
    r = subprocess.run(
        [bin_, "--headless", "--norestore", path,
         "macro:///Standard.Module1.RecalculateAndSave"],
        capture_output=True, timeout=180)
    if r.returncode != 0:
        sys.exit(f"LibreOffice falló: {r.stderr.decode()[:400]}")
    # verificación rápida: alguna celda de fórmula debe tener valor cacheado
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else wb.active
    print(f"Recalculado OK: {path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1])
