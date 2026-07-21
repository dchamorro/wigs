#!/usr/bin/env python3
"""Hero screen del WIG de compañía para TRMNL (usetrmnl.com) — extrae y publica.

Lee el Dashboard de un tablero YA RECALCULADO (scripts/recalc.py: openpyxl solo
ve valores cacheados) y produce:

  dist/trmnl/hero.svg            — render local con los datos reales (revisión)
  dist/trmnl/hero.png            — ídem rasterizado (--png, requiere cairosvg)
  dist/trmnl/hero_markup.liquid  — markup del private plugin de TRMNL: el MISMO
                                   template de render_hero() pero con
                                   placeholders Liquid; se pega UNA vez en
                                   TRMNL → Private Plugin → Edit Markup
  dist/trmnl/hero_payload.json   — merge variables listos para el webhook

Con --push además publica: POST {"merge_variables": …} al webhook del plugin
(https://usetrmnl.com/api/custom_plugins/<uuid>), que TRMNL renderiza en su
servidor y empuja al e-ink en el siguiente refresh. La URL sale de
--webhook-url, o de las env TRMNL_WEBHOOK_URL / TRMNL_PLUGIN_UUID.

Uso:
  python3 scripts/trmnl_hero.py --xlsx dist/demo_data.xlsx --out dist/trmnl
  TRMNL_WEBHOOK_URL=… python3 scripts/trmnl_hero.py --xlsx web/tablero.xlsx --push

Setup completo del plugin: docs/TRMNL.md.
"""
import argparse
import datetime
import json
import os
import sys
import urllib.request

try:
    from . import trmnl_render as tr
except ImportError:
    import trmnl_render as tr

MESES = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]
DIAS = ["lun", "mar", "mié", "jue", "vie", "sáb", "dom"]
ST_KEY = {"en meta": "meta", "riesgo": "riesgo", "atrasado": "atrasado"}
BAR_W = 420          # ancho de las barras del hero (px)
PAYLOAD_MAX = 2048   # límite del webhook de TRMNL (2KB)


def fmt_money(v):
    if v is None:
        return "—"
    a = abs(v)
    if a >= 1_000_000:
        s = f"${v / 1_000_000:.1f}M"
        return s.replace(".0M", "M")
    if a >= 1_000:
        return f"${v / 1_000:,.0f}K"
    return f"${v:,.0f}"


def shorten(s, n=34):
    s = " ".join(str(s or "").split())
    if len(s) <= n:
        return s
    cut = s[:n].rsplit(" ", 1)[0]
    return cut + "…"


def st_key(estado):
    return ST_KEY.get(str(estado or "").strip().lower(), "none")


def extract_hero(xlsx_path, today=None):
    """Dashboard → view-model plano del hero (posiciones del CONTRATO)."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Dashboard"]
    today = today or datetime.date.today()

    meta_anual = ws["C19"].value or 0
    year = ws["A22"].value.year if isinstance(ws["A22"].value, datetime.datetime) else today.year

    # seguimiento mensual (filas 22–33): último mes con real digitado
    last = None
    for r in range(22, 34):
        if ws.cell(row=r, column=3).value is not None:
            last = r
    if last:
        mes = ws.cell(row=last, column=1).value.month
        acum_meta = ws.cell(row=last, column=4).value or 0
        acum_real = ws.cell(row=last, column=5).value or 0
        ritmo_pct = ws.cell(row=last, column=6).value
        estado = ws.cell(row=last, column=7).value
        ritmo_label = f'{estado} al corte de {MESES[mes - 1]}'
        ritmo_note = (f'Acumulado a {MESES[mes - 1]}: {fmt_money(acum_real)} de '
                      f'{fmt_money(acum_meta)} del plan ({round((ritmo_pct or 0) * 100)}%)')
    else:
        acum_real, estado = 0, None
        ritmo_label, ritmo_note = "Sin datos aún", "Aún no hay meses con NAT real digitado"

    # cobertura de backlog (filas 14–16): la fila del año en curso
    backlog_pct, backlog_note = 0, "—"
    for r in range(14, 17):
        if ws.cell(row=r, column=1).value == year:
            gp_meta = ws.cell(row=r, column=2).value or 0
            comprometido = ws.cell(row=r, column=3).value or 0
            brecha = ws.cell(row=r, column=4).value or 0
            backlog_pct = round((ws.cell(row=r, column=5).value or 0) * 100)
            backlog_note = (f'GP comprometido {fmt_money(comprometido)} de '
                            f'{fmt_money(gp_meta)} · brecha {fmt_money(brecha)}')

    # WIGs de soporte (filas 37–48): # · nombre · estado
    wigs, counts = [], {"meta": 0, "riesgo": 0, "atrasado": 0, "none": 0}
    for r in range(37, 49):
        num = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        if num is None or name is None:
            continue
        key = st_key(ws.cell(row=r, column=7).value)
        counts[key] += 1
        wigs.append({"label": f'{num} · {shorten(name)}', "st": key})

    nat_pct = round(100 * acum_real / meta_anual) if meta_anual else 0
    date_label = f'{DIAS[today.weekday()]} {today.day} {MESES[today.month - 1]} {today.year}'
    return {
        "title": f'WIG de Compañía · {fmt_money(meta_anual)} NAT {year}',
        "subtitle": "GBM Nicaragua · 4 Disciplinas de Ejecución",
        "week_label": f'SEMANA {today.isocalendar()[1]}',
        "date_label": date_label,
        "updated_label": f'Actualizado {date_label} · fuente: tablero WIG',
        "nat_year": year,
        "nat_meta": fmt_money(meta_anual),
        "nat_real": fmt_money(acum_real),
        "nat_pct": nat_pct,
        "nat_bar_w": round(BAR_W * min(1, nat_pct / 100)),
        "nat_st": st_key(estado),
        "ritmo_label": ritmo_label,
        "ritmo_note": ritmo_note,
        "backlog_pct": backlog_pct,
        "backlog_bar_w": round(BAR_W * min(1, backlog_pct / 100)),
        "backlog_note": backlog_note,
        "n_meta": counts["meta"],
        "n_riesgo": counts["riesgo"],
        "n_atrasado": counts["atrasado"] + counts["none"],
        "wigs": wigs,
    }


# ── markup Liquid (se pega una sola vez en TRMNL) ─────────────────────────────
SCALAR_KEYS = ["title", "subtitle", "week_label", "date_label", "updated_label",
               "nat_year", "nat_meta", "nat_real", "nat_pct", "nat_bar_w", "nat_st",
               "ritmo_label", "ritmo_note", "backlog_pct", "backlog_bar_w",
               "backlog_note", "n_meta", "n_riesgo", "n_atrasado"]

def placeholder_vm(n_wigs=12):
    vm = {k: f'{{{{ {k} }}}}' for k in SCALAR_KEYS}
    vm["wigs"] = [{"label": f'{{{{ w{i} }}}}', "st": f'{{{{ s{i} }}}}'}
                  for i in range(1, n_wigs + 1)]
    return vm

def build_markup():
    svg = tr.render_hero({"hero": placeholder_vm()})
    return ("<!-- Marcador WIG — hero. Pegar TODO este archivo en\n"
            "     usetrmnl.com → Plugins → Private Plugin → Edit Markup.\n"
            "     Los {{ … }} llegan por webhook (scripts/trmnl_hero.py --push). -->\n"
            + svg)

def to_payload(vm):
    """View-model → merge variables planos del webhook (w1/s1 … w12/s12)."""
    p = {k: vm[k] for k in SCALAR_KEYS}
    for i, wig in enumerate(vm["wigs"], 1):
        p[f'w{i}'] = wig["label"]
        p[f's{i}'] = wig["st"]
    return p


# ── publicación ───────────────────────────────────────────────────────────────
def webhook_url(args):
    if args.webhook_url:
        return args.webhook_url
    if os.environ.get("TRMNL_WEBHOOK_URL"):
        return os.environ["TRMNL_WEBHOOK_URL"]
    if os.environ.get("TRMNL_PLUGIN_UUID"):
        return f'https://usetrmnl.com/api/custom_plugins/{os.environ["TRMNL_PLUGIN_UUID"]}'
    return None

def push(payload, url):
    body = json.dumps({"merge_variables": payload}, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=body,
                                 headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.status, resp.read().decode("utf-8", "replace")


def main(argv=None):
    ap = argparse.ArgumentParser(description="Hero WIG para TRMNL: extrae del tablero y publica.")
    ap.add_argument("--xlsx", required=True, help="tablero recalculado (dist/demo_data.xlsx, web/tablero.xlsx…)")
    ap.add_argument("--out", default="dist/trmnl", help="carpeta de salida")
    ap.add_argument("--png", action="store_true", help="rasterizar hero.png (cairosvg)")
    ap.add_argument("--push", action="store_true", help="publicar merge variables al webhook de TRMNL")
    ap.add_argument("--webhook-url", help="URL del webhook (o env TRMNL_WEBHOOK_URL / TRMNL_PLUGIN_UUID)")
    args = ap.parse_args(argv)

    vm = extract_hero(args.xlsx)
    payload = to_payload(vm)
    os.makedirs(args.out, exist_ok=True)

    svg = tr.render_hero({"hero": vm})
    with open(os.path.join(args.out, "hero.svg"), "w", encoding="utf-8") as f:
        f.write(svg)
    print(f'✓ {args.out}/hero.svg')
    if args.png:
        if tr.to_png(svg, os.path.join(args.out, "hero.png")):
            print(f'✓ {args.out}/hero.png')
        else:
            print("⚠ cairosvg no instalado — solo SVG.", file=sys.stderr)

    with open(os.path.join(args.out, "hero_markup.liquid"), "w", encoding="utf-8") as f:
        f.write(build_markup())
    print(f'✓ {args.out}/hero_markup.liquid  (pegar una vez en el plugin de TRMNL)')

    body = json.dumps(payload, ensure_ascii=False, indent=2)
    with open(os.path.join(args.out, "hero_payload.json"), "w", encoding="utf-8") as f:
        f.write(body + "\n")
    size = len(json.dumps({"merge_variables": payload}, ensure_ascii=False).encode("utf-8"))
    print(f'✓ {args.out}/hero_payload.json  ({size} bytes de {PAYLOAD_MAX} del webhook)')
    if size > PAYLOAD_MAX:
        print(f'✗ payload excede el límite de {PAYLOAD_MAX} bytes del webhook de TRMNL', file=sys.stderr)
        return 1

    if args.push:
        url = webhook_url(args)
        if not url:
            print("✗ --push sin webhook: usar --webhook-url, TRMNL_WEBHOOK_URL o "
                  "TRMNL_PLUGIN_UUID (ver docs/TRMNL.md)", file=sys.stderr)
            return 1
        status, resp = push(payload, url)
        print(f'✓ publicado a TRMNL ({status}): {resp.strip()}')
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
