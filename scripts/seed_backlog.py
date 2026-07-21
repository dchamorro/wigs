#!/usr/bin/env python3
"""Siembra el grid semanal de `Backlog Readings` en Airtable desde el tablero.

Lee las páginas por año (2027/2028/2029) de web/tablero.xlsx — col A (semana)
desde la fila 12 — y crea en Airtable una fila por lunes por año, para que el
equipo solo digite `GP comprometido` en el grid cada semana y para que
`trmnl_airtable` derive el número de semana del grid.

Idempotente: las parejas (año, fecha) ya existentes se saltan; correrlo dos
veces no duplica. Si el Excel trae GP comprometido (col B), lo copia.

Uso:
  AIRTABLE_PAT=pat… python3 scripts/seed_backlog.py [--xlsx web/tablero.xlsx] [--dry-run]

El PAT necesita data.records:write sobre la base WIGS (usar el PAT personal,
NO el de solo lectura del CI).
"""
import argparse
import json
import os
import sys
import time
import urllib.request
from datetime import date, datetime

try:
    import trmnl_airtable as ta
except ImportError:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from scripts import trmnl_airtable as ta

DATA0 = 12  # primera fila de la tabla semanal en las páginas por año


def post_records(table_id, token, rows, base_id=ta.BASE_ID):
    """POST en lotes de 10 (límite de Airtable), con pausa por rate limit."""
    url = f"https://api.airtable.com/v0/{base_id}/{table_id}"
    for i in range(0, len(rows), 10):
        body = json.dumps({"records": rows[i:i + 10]}).encode("utf-8")
        req = urllib.request.Request(url, data=body, method="POST", headers={
            "Authorization": f"Bearer {token}", "Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            json.load(resp)
        time.sleep(0.25)


def weeks_from_xlsx(path):
    """{año: [(fecha, gp|None), …]} desde las páginas por año del tablero."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    years = {}
    for name in wb.sheetnames:
        if not (len(name) == 4 and name.isdigit()):
            continue
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=DATA0, min_col=1, max_col=2):
            a, b = row[0].value, row[1].value
            if not isinstance(a, (datetime, date)):
                break
            fecha = a.date() if isinstance(a, datetime) else a
            rows.append((fecha, b if isinstance(b, (int, float)) else None))
        years[name] = rows
    wb.close()
    return years


def main(argv=None):
    ap = argparse.ArgumentParser(description="Siembra Backlog Readings desde el tablero.")
    ap.add_argument("--xlsx", default="web/tablero.xlsx")
    ap.add_argument("--dry-run", action="store_true", help="mostrar el plan sin escribir")
    args = ap.parse_args(argv)

    token = os.environ.get("AIRTABLE_PAT")
    if not token:
        raise SystemExit("Falta AIRTABLE_PAT (con permiso de escritura sobre la base WIGS)")

    grid = weeks_from_xlsx(args.xlsx)
    year_ids = {r["fields"].get("Año"): r["id"] for r in ta.fetch_table(ta.T_YEARS, token)}
    existing = {(r["fields"]["Año"][0], r["fields"]["Fecha"])
                for r in ta.fetch_table(ta.T_BACKLOG, token)
                if r.get("fields", {}).get("Año") and r["fields"].get("Fecha")}

    to_create = []
    for year, rows in sorted(grid.items()):
        rec_id = year_ids.get(year)
        if not rec_id:
            print(f"⚠ {year}: sin fila en Years — se salta", file=sys.stderr)
            continue
        nuevos = 0
        for i, (fecha, gp) in enumerate(rows, start=1):
            if (rec_id, fecha.isoformat()) in existing:
                continue
            fields = {"Etiqueta": f"{year} · sem {i:02d} · {fecha.day}-{ta.MESES[fecha.month - 1]}",
                      "Fecha": fecha.isoformat(), "Año": [rec_id]}
            if gp is not None:
                fields["GP comprometido"] = gp
            to_create.append({"fields": fields})
            nuevos += 1
        rango = f"{rows[0][0]:%d-%b} → {rows[-1][0]:%d-%b}" if rows else "vacío"
        print(f"{year}: {len(rows)} semanas en el tablero ({rango}) · "
              f"{nuevos} nuevas · {len(rows) - nuevos} ya existían")

    if not to_create:
        print("Nada que crear — el grid ya está sembrado.")
        return 0
    if args.dry_run:
        print(f"[dry-run] Se crearían {len(to_create)} filas en Backlog Readings.")
        return 0
    post_records(ta.T_BACKLOG, token, to_create)
    print(f"✓ {len(to_create)} filas creadas en Backlog Readings.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
