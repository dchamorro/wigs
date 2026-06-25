"""
sync_from_airtable.py — genera el Excel DESDE Airtable (Opción A de la arquitectura).

Airtable es la fuente de verdad; el Excel pasa a ser un **artefacto derivado**.
Este módulo lee las ENTRADAS de la base WIGS y las escribe en un tablero recién
construido por `build_wig.py` — exactamente las mismas celdas azules que copia
`migrate_data.py`, emparejando periodos por FECHA (no por posición), de modo que
cambiar el rango de semanas no desalinea nada.

Dos fuentes de datos (misma forma de "bundle", igual que scripts/trmnl_render.py):
  - from_airtable(base_id, token) -> bundle   (producción; API REST, corre en CI)
  - from_fixture(path)            -> bundle   (pruebas/offline; JSON)

Flujo:
  python3 excel/build_wig.py            dist/Tablero_WIG_4DX_GBM.xlsx     # estructura en blanco
  python3 excel/sync_from_airtable.py   dist/Tablero_WIG_4DX_GBM.xlsx OUT.xlsx
  python3 scripts/recalc.py             OUT.xlsx                          # OBLIGATORIO

  # offline / pruebas (sin red):
  python3 excel/sync_from_airtable.py NEW.xlsx OUT.xlsx --fixture excel/airtable_sample.json

Env para la fuente Airtable: AIRTABLE_TOKEN (PAT) y AIRTABLE_BASE_ID.

NO calcula nada: solo coloca entradas. El cálculo (acum., %, estado, cobertura)
lo hace LibreOffice en `recalc.py`. El contrato Excel↔parser queda intacto.
"""
import os
import re
import sys
import json
from datetime import date

from openpyxl import load_workbook

# Reutiliza los helpers ya probados de migrate_data (mismo directorio).
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from migrate_data import is_year_page, date_map, _find_annual_nat, _find_key_rows  # noqa: E402

MAX_LEADS = 8
LEAD_COL = lambda slot: 9 + 2 * (int(slot) - 1)   # slot 1->I(9), 2->K(11), ...


# ───────────────────────── normalización ────────────────────────────────────
def _d(v):
    """A `date` desde 'YYYY-MM-DD' | date | datetime | None."""
    if v is None:
        return None
    if isinstance(v, str):
        return date.fromisoformat(v[:10])
    return v.date() if hasattr(v, 'date') else v


# ───────────────────────── fuente: fixture JSON ─────────────────────────────
def from_fixture(path):
    with open(path, encoding='utf-8') as f:
        return json.load(f)


# ───────────────────────── fuente: Airtable REST ────────────────────────────
def from_airtable(base_id, token):
    """Construye el bundle leyendo la base vía API REST de Airtable.

    Tablas esperadas (por nombre): WIGs, Leads, Captura semanal, Years,
    Backlog Readings, NAT Monthly. (Solo se usa en CI; aquí no hay egress.)
    """
    import requests
    s = requests.Session()
    s.headers.update({'Authorization': f'Bearer {token}'})

    def schema():
        r = s.get(f'https://api.airtable.com/v0/meta/bases/{base_id}/tables')
        r.raise_for_status()
        return {t['name']: t['id'] for t in r.json()['tables']}

    def records(tid):
        out, offset = [], None
        while True:
            params = {'pageSize': 100}
            if offset:
                params['offset'] = offset
            r = s.get(f'https://api.airtable.com/v0/{base_id}/{tid}', params=params)
            r.raise_for_status()
            j = r.json()
            out += j['records']
            offset = j.get('offset')
            if not offset:
                return out

    tid = schema()
    g = lambda rec, k, d=None: rec['fields'].get(k, d)

    # WIGs: id -> nombre/tab; dueño y meta
    wig_by_id, wigs = {}, {}
    for rec in records(tid['WIGs']):
        name = g(rec, 'Nombre')
        wig_by_id[rec['id']] = name
        wigs[name] = {'dueno': g(rec, 'Dueño'), 'meta': g(rec, 'Meta'),
                      'leads': {}, 'readings': []}

    # Leads: nombre y meta por slot, dentro de su WIG
    for rec in records(tid['Leads']):
        link = g(rec, 'WIG') or []
        if not link:
            continue
        name = wig_by_id.get(link[0])
        if name not in wigs:
            continue
        slot = g(rec, 'Slot')
        if slot:
            wigs[name]['leads'][str(int(slot))] = {'name': g(rec, 'Nombre'), 'meta': g(rec, 'Meta')}

    # Captura semanal (ancha): 1 fila por WIG por semana → lag + L1..L8
    for rec in records(tid.get('Captura semanal', '')) if 'Captura semanal' in tid else []:
        link = g(rec, 'WIG') or []
        name = wig_by_id.get(link[0]) if link else None
        if name not in wigs:
            continue
        leads = {str(k): g(rec, f'L{k}') for k in range(1, MAX_LEADS + 1) if g(rec, f'L{k}') is not None}
        wigs[name]['readings'].append({'date': g(rec, 'Fecha'), 'lag': g(rec, 'Lag real'), 'leads': leads})

    # Years + Backlog + NAT
    years, traj = {}, {}
    year_by_id = {}
    for rec in records(tid['Years']):
        y = str(g(rec, 'Año'))
        year_by_id[rec['id']] = y
        years[y] = {'gp_pct_assumed': g(rec, 'GP %') if g(rec, 'GP% supuesto') else None,
                    'backlog': []}
        if g(rec, 'NAT real') is not None:
            traj[y] = g(rec, 'NAT real')
    for rec in records(tid.get('Backlog Readings', '')) if 'Backlog Readings' in tid else []:
        link = g(rec, 'Año') or []
        y = year_by_id.get(link[0]) if link else None
        if y in years:
            years[y]['backlog'].append({'date': g(rec, 'Fecha'), 'gp_committed': g(rec, 'GP comprometido')})

    nat_monthly = []
    for rec in records(tid.get('NAT Monthly', '')) if 'NAT Monthly' in tid else []:
        if g(rec, 'NAT real') is not None:
            nat_monthly.append({'date': g(rec, 'Fecha'), 'nat_real': g(rec, 'NAT real')})

    return {'wigs': wigs, 'years': years, 'nat_trajectory': traj,
            'nat_annual_meta': None, 'nat_monthly': nat_monthly}


# ───────────────────────── escritura al workbook ────────────────────────────
def _set(ws, row, col, v):
    """Escribe v si no es None (nunca pisa con vacío; nunca escribe fórmula)."""
    if v is None:
        return 0
    ws.cell(row=row, column=col, value=v)
    return 1


def apply_to_workbook(bundle, new_path, out_path):
    """Vuelca las entradas del bundle en un tablero recién construido."""
    wb = load_workbook(new_path, data_only=False)
    report = []

    # ── Pestañas de WIG ──
    for name, w in (bundle.get('wigs') or {}).items():
        if name not in wb.sheetnames:
            report.append(f'  {name}: no existe en el tablero — omitido')
            continue
        ws = wb[name]
        n = _set(ws, 4, 2, w.get('dueno')) + _set(ws, 5, 2, w.get('meta'))
        for slot, ld in (w.get('leads') or {}).items():
            c = LEAD_COL(slot)
            n += _set(ws, 8, c, ld.get('name')) + _set(ws, 9, c, ld.get('meta'))
        nmap = date_map(ws)
        moved = 0
        for rd in (w.get('readings') or []):
            row = nmap.get(_d(rd.get('date')))
            if row is None:
                continue
            moved += _set(ws, row, 4, rd.get('lag'))
            for slot, val in (rd.get('leads') or {}).items():
                moved += _set(ws, row, LEAD_COL(slot), val)
        report.append(f'  {name}: {n} de cabecera/leads, {moved} de lecturas')

    # ── Páginas por año (backlog GP por semana + GP% supuesto) ──
    for y, yd in (bundle.get('years') or {}).items():
        if y not in wb.sheetnames:
            continue
        ws = wb[y]
        if isinstance(yd.get('gp_pct_assumed'), (int, float)):
            ws['E4'] = yd['gp_pct_assumed']
        amap, r = {}, 12
        while ws.cell(row=r, column=1).value is not None:
            amap[_d(ws.cell(row=r, column=1).value)] = r
            r += 1
        moved = 0
        for b in (yd.get('backlog') or []):
            row = amap.get(_d(b.get('date')))
            if row:
                moved += _set(ws, row, 2, b.get('gp_committed'))
        report.append(f'  {y}: {moved} semanas de backlog')

    # ── Dashboard: meta anual NAT, NAT real mensual y NAT real por año ──
    if 'Dashboard' in wb.sheetnames:
        d = wb['Dashboard']
        a = _find_annual_nat(d)
        if a and bundle.get('nat_annual_meta') is not None:
            d.cell(row=a[0], column=a[1], value=bundle['nat_annual_meta'])
        meses = _find_key_rows(d, 'Mes')
        for m in (bundle.get('nat_monthly') or []):
            row = meses.get(_d(m.get('date')))
            if row:
                _set(d, row, 3, m.get('nat_real'))
        # trayectoria: las etiquetas son '2027 (meta)' → empareja por los 4 dígitos
        yrows = {}
        for label, row in _find_key_rows(d, 'Año').items():
            mo = re.search(r'\d{4}', str(label))
            if mo:
                yrows[mo.group(0)] = row
        for y, val in (bundle.get('nat_trajectory') or {}).items():
            row = yrows.get(str(y))
            if row:
                _set(d, row, 6, val)
        report.append('  Dashboard: NAT (meta anual, mensual, trayectoria)')

    wb.save(out_path)
    print(f'Sync Airtable → {out_path}')
    print('\n'.join(report))
    print('\nIMPORTANTE: recalcular antes de publicar:')
    print(f'  python3 scripts/recalc.py {out_path}')


# ───────────────────────── CLI ──────────────────────────────────────────────
def main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 1
    new_path, out_path = argv[0], argv[1]
    fixture = None
    if '--fixture' in argv:
        fixture = argv[argv.index('--fixture') + 1]

    if fixture:
        bundle = from_fixture(fixture)
    else:
        token = os.environ.get('AIRTABLE_TOKEN')
        base = os.environ.get('AIRTABLE_BASE_ID')
        if not token or not base:
            sys.exit('Falta AIRTABLE_TOKEN / AIRTABLE_BASE_ID (o usa --fixture ruta.json)')
        bundle = from_airtable(base, token)

    apply_to_workbook(bundle, new_path, out_path)
    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
