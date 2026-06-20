"""
fill_demo_data.py — Llena un tablero recién construido con datos de demostración.
Uso: python3 excel/fill_demo_data.py ENTRADA.xlsx SALIDA.xlsx
Después recalcular: python3 scripts/recalc.py SALIDA.xlsx
"""
import sys, random, shutil
from openpyxl import load_workbook

PROFILES = {
 '1. Smart User 500':  ('R. Castillo', [15,22,18,24,16,20,25,19,21,17,23,20], 1.05),
 '2. MV Advisory':     ('M. Aguilar',  [180,220,150,250,200,160,240,210,190,230,200,210], 0.88),
 '3. Negocios +30K GP':('J. Roque',    [0,0,0,0,0,0,0,1,0,0,0,0], 0.75),
 '4. Entregas 100%':   ('K. Solis',    [1,.96,1,1,.95,1,1,.98,1,1,1,1], 1.0),
 '5. MV Datacenter':   ('A. Vargas',   [0,0,1800,0,950,0,0,1200,0,600,0,800], 1.1),
 '6. Cero Expired':    ('D. Lacayo',   [5,4,4,3,2,2,1,1,0,0,1,0], 0.95),
 '7. MV Digital Sol':  ('S. Morales',  [0,0,2500,0,0,1800,0,0,2200,0,0,0], 0.85),
 '8. GP 16%':          ('F. Tellez',   [.148,.152,.156,.161,.163], 1.0),
 '9. CSAT':            ('L. Espinoza', [4.1,4.2,4.3,4.4,4.5], 0.9),
 '10. Apalancamiento': ('F. Tellez',   [.13,.125,.118,.11,.10], 0.95),
 '11. Recuperar DS':   ('S. Morales',  [-78000,-62000,-45000,-28000,-12000], 0.9),
 '12. Utilizacion':    ('A. Vargas',   [.42,.435,.45,.46,.47], 0.92),
}

def main(inp, out):
    shutil.copy(inp, out)
    wb = load_workbook(out)
    random.seed(11)
    for tab,(duenio,reals,bias) in PROFILES.items():
        if tab not in wb.sheetnames: continue
        s = wb[tab]; s['B4']=duenio
        leads=[]
        for k in range(8):
            c=9+2*k
            nm=s.cell(row=8,column=c).value
            mt=s.cell(row=9,column=c).value
            if nm and '(Disponible)' not in str(nm) and mt is not None:
                leads.append((c,mt))
        for i,re in enumerate(reals):
            r=11+i
            s.cell(row=r,column=4,value=re)
            for c,mt in leads:
                mult=max(0.3,min(1.6,random.gauss(bias,0.18)))
                v=mt*mult
                if mt==1.0 and isinstance(mt,float): v=min(1.0,v)
                if isinstance(mt,int) and mt<60: v=round(v)
                else: v=round(v,2)
                s.cell(row=r,column=c,value=v)
    d = wb['Dashboard']
    # NAT real mensual del año en curso: bloque mensual desde la fila 22 (col C)
    for i,v in enumerate([70000,78000,90000,85000,95000]):
        d.cell(row=22+i,column=3,value=v)
    # Backlog: GP comprometido acumulado por semana (col B, desde fila 12) — demo
    BL_DEMO = {
        'Backlog 2027': [2000000,2150000,2300000,2450000,2600000],
        'Backlog 2028': [1600000,1900000,2150000,2350000,2500000],
        'Backlog 2029': [700000,950000,1150000,1350000,1550000],
    }
    for tab,vals in BL_DEMO.items():
        if tab not in wb.sheetnames: continue
        bs = wb[tab]
        for i,v in enumerate(vals):
            bs.cell(row=12+i,column=2,value=v)
    if 'Compromisos' in wb.sheetnames:
        from datetime import date, timedelta
        cs = wb['Compromisos']
        lunes = [date(2026,6,15)+timedelta(days=7*i) for i in range(8)]
        demo = [
            (lunes[6],1,1,'Firmar asignación de 20 equipos con Promerica','R. Castillo','Hecho'),
            (lunes[6],1,2,'Agendar demo Smart User con Disnorte','C. Mendoza','Hecho'),
            (lunes[7],1,1,'Cerrar compromiso de asignación con BAC (15 equipos)','R. Castillo','Pendiente'),
            (lunes[7],1,2,'Preparar POC para cliente retail','C. Mendoza','Pendiente'),
            (lunes[5],2,1,'Presentar propuesta advisory a Lafise','M. Aguilar','Hecho'),
            (lunes[6],2,1,'Enviar propuesta de health check a Claro','M. Aguilar','Hecho'),
            (lunes[7],2,2,'Cotizar 40 horas advisory para banco regional','P. Downs','Pendiente'),
            (lunes[6],4,1,'Completar checklist de entregas de la semana','K. Solis','Hecho'),
            (lunes[7],4,2,'Escalar riesgo de inventario del proyecto X','K. Solis','Hecho'),
            (lunes[7],6,1,'Plan de renovación para 3 contratos a 90 días','D. Lacayo','Pendiente'),
        ]
        for i,(f,wg,ld,txt,resp,est) in enumerate(demo):
            r = 2+i
            cs.cell(row=r,column=1,value=f)
            cs.cell(row=r,column=2,value=wg)
            cs.cell(row=r,column=3,value=ld)
            cs.cell(row=r,column=4,value=txt)
            cs.cell(row=r,column=5,value=resp)
            cs.cell(row=r,column=6,value=est)
    wb.save(out)
    print('demo data ->', out)

if __name__ == '__main__':
    if len(sys.argv)!=3: print(__doc__); sys.exit(1)
    main(sys.argv[1], sys.argv[2])
