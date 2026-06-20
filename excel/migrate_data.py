"""
migrate_data.py — Migra los datos digitados por los dueños desde un tablero
existente hacia un tablero recién construido (tras un cambio estructural).

Uso:
    python3 excel/migrate_data.py VIEJO.xlsx NUEVO.xlsx SALIDA.xlsx

Copia (solo celdas de entrada, nunca fórmulas):
  - Por pestaña de WIG: dueño (B4), meta (B5), nombres de leads (fila 8),
    metas de leads (fila 9), columna Real del lag (D) y columnas Real de
    cada lead — emparejando filas por la FECHA de la columna B, de modo
    que un cambio en el rango de periodos no desalinee los datos.
  - Dashboard: meta anual NAT, NAT real mensual (emparejado por mes) y NAT
    real anual de la trayectoria (emparejado por año). Se localizan por
    etiqueta, así sirve aunque viejo y nuevo tengan el Dashboard en filas
    distintas (p. ej. la reestructuración a 12 WIGs movió el bloque mensual).

Después de migrar, recalcular:  python3 scripts/recalc.py SALIDA.xlsx
"""
import sys
from openpyxl import load_workbook

DATA0 = 11
MAX_LEADS = 8
SKIP = {'Dashboard', 'Instrucciones', 'Compromisos'}


def date_map(ws):
    """fila -> fecha (col B) para emparejar periodos entre versiones."""
    out = {}
    r = DATA0
    while ws.cell(row=r, column=2).value is not None:
        v = ws.cell(row=r, column=2).value
        key = v.date() if hasattr(v, 'date') else v
        out[key] = r
        r += 1
    return out


def _find_annual_nat(ws):
    """Devuelve (fila, col) de la celda de meta anual NAT (a 2 columnas a la
    derecha de la etiqueta 'Meta anual NAT:'), o None si no está."""
    for r in range(1, 60):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().startswith('Meta anual NAT'):
            return (r, 3)
    return None


def _find_key_rows(ws, header):
    """Localiza el encabezado `header` en la columna A y devuelve {clave: fila}
    para las filas de datos debajo, hasta la primera fila sin etiqueta en A.
    Sirve para los bloques 'Mes' (mensual NAT) y 'Año' (trayectoria)."""
    hdr = None
    for r in range(1, 60):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == header:
            hdr = r
            break
    if hdr is None:
        return {}
    out, r = {}, hdr + 1
    while True:
        v = ws.cell(row=r, column=1).value
        if v is None or (isinstance(v, str) and v.strip() == ''):
            break
        key = v.date() if hasattr(v, 'date') else v
        out[key] = r
        r += 1
    return out


def main(old_path, new_path, out_path):
    old = load_workbook(old_path, data_only=False)
    new = load_workbook(new_path, data_only=False)
    report = []

    for name in new.sheetnames:
        if name in SKIP or name not in old.sheetnames:
            if name not in old.sheetnames and name not in SKIP:
                report.append(f'  {name}: pestaña nueva, sin datos que migrar')
            continue
        o, n = old[name], new[name]
        # encabezado
        for a in ('B4', 'B5'):
            if o[a].value is not None:
                n[a] = o[a].value
        # nombres y metas de leads (si el viejo los tenía personalizados)
        for k in range(MAX_LEADS):
            c = 9 + 2 * k
            for row in (8, 9):
                v = o.cell(row=row, column=c).value
                if v is not None:
                    n.cell(row=row, column=c, value=v)
        # datos por fecha
        omap, nmap = date_map(o), date_map(n)
        moved = 0
        for fecha, orow in omap.items():
            nrow = nmap.get(fecha)
            if nrow is None:
                continue
            cols = [4] + [9 + 2 * k for k in range(MAX_LEADS)]
            for c in cols:
                v = o.cell(row=orow, column=c).value
                if v is not None and not (isinstance(v, str) and v.startswith('=')):
                    n.cell(row=nrow, column=c, value=v)
                    moved += 1
        dropped = len(set(omap) - set(nmap))
        report.append(f'  {name}: {moved} celdas migradas' + (f' · ⚠ {dropped} periodos del viejo no existen en el nuevo' if dropped else ''))

    # Compromisos: copia plana de filas digitadas (A2:F...)
    if 'Compromisos' in old.sheetnames and 'Compromisos' in new.sheetnames:
        oc, nc = old['Compromisos'], new['Compromisos']
        moved = 0
        for r in range(2, oc.max_row + 1):
            if all(oc.cell(row=r, column=c).value is None for c in range(1, 7)):
                continue
            for c in range(1, 7):
                v = oc.cell(row=r, column=c).value
                if v is not None and not (isinstance(v, str) and v.startswith('=')):
                    nc.cell(row=r, column=c, value=v)
            moved += 1
        report.append(f'  Compromisos: {moved} filas migradas')

    # Dashboard NAT — se localiza por etiqueta (robusto ante cambios de fila)
    if 'Dashboard' in old.sheetnames:
        od, nd = old['Dashboard'], new['Dashboard']
        # meta anual NAT: celda a la derecha de la etiqueta 'Meta anual NAT:'
        oa, na = _find_annual_nat(od), _find_annual_nat(nd)
        if oa is not None and na is not None and od.cell(row=oa[0], column=oa[1]).value is not None:
            nd.cell(row=na[0], column=na[1], value=od.cell(row=oa[0], column=oa[1]).value)
        # NAT real mensual (col C bajo el encabezado 'Mes'), emparejado por mes
        omeses, nmeses = _find_key_rows(od, 'Mes'), _find_key_rows(nd, 'Mes')
        for mes, orow in omeses.items():
            nrow = nmeses.get(mes)
            if nrow:
                v = od.cell(row=orow, column=3).value
                if v is not None and not (isinstance(v, str) and v.startswith('=')):
                    nd.cell(row=nrow, column=3, value=v)
        # NAT real anual de la trayectoria (col F bajo el encabezado 'Año'), por año
        oyrs, nyrs = _find_key_rows(od, 'Año'), _find_key_rows(nd, 'Año')
        for yr, orow in oyrs.items():
            nrow = nyrs.get(yr)
            if nrow:
                v = od.cell(row=orow, column=6).value
                if v is not None and not (isinstance(v, str) and v.startswith('=')):
                    nd.cell(row=nrow, column=6, value=v)
        report.append('  Dashboard: NAT migrado')

    new.save(out_path)
    print(f'Migrado {old_path} -> {out_path}')
    print('\n'.join(report))
    print('\nIMPORTANTE: recalcular antes de publicar:')
    print(f'  python3 scripts/recalc.py {out_path}')


if __name__ == '__main__':
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1], sys.argv[2], sys.argv[3])
