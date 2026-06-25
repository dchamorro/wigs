from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as col
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, timedelta

ARIAL = 'Arial'
F_TITLE = Font(name=ARIAL, size=14, bold=True, color='1F3864')
F_SECT = Font(name=ARIAL, size=12, bold=True, color='1F3864')
F_LBL = Font(name=ARIAL, size=10, bold=True)
F_TXT = Font(name=ARIAL, size=10)
F_INPUT = Font(name=ARIAL, size=10, color='0000FF')
F_INPUT_S = Font(name=ARIAL, size=9, color='0000FF')
F_FORM = Font(name=ARIAL, size=10, color='000000')
F_LINK = Font(name=ARIAL, size=10, color='008000')
F_HDR = Font(name=ARIAL, size=10, bold=True, color='FFFFFF')
F_HDR_S = Font(name=ARIAL, size=9, bold=True, color='FFFFFF')
FILL_HDR = PatternFill('solid', start_color='1F3864')
FILL_LEAD = PatternFill('solid', start_color='2E5496')
FILL_YEL = PatternFill('solid', start_color='FFFF00')
FILL_GREY = PatternFill('solid', start_color='F2F2F2')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')
WRAPC = Alignment(horizontal='center', vertical='center', wrap_text=True)
WRAP = Alignment(vertical='top', wrap_text=True)

GREEN_DXF = DifferentialStyle(fill=PatternFill('solid', start_color='C6EFCE'), font=Font(color='006100', name=ARIAL))
YEL_DXF = DifferentialStyle(fill=PatternFill('solid', start_color='FFEB9C'), font=Font(color='9C6500', name=ARIAL))
RED_DXF = DifferentialStyle(fill=PatternFill('solid', start_color='FFC7CE'), font=Font(color='9C0006', name=ARIAL))

def estado_cf(ws, rng):
    first = rng.split(':')[0]
    for txt, dxf in (('En meta', GREEN_DXF), ('Riesgo', YEL_DXF), ('Atrasado', RED_DXF)):
        ws.conditional_formatting.add(rng, Rule(type='containsText', operator='containsText', text=txt, dxf=dxf, formula=[f'NOT(ISERROR(SEARCH("{txt}",{first})))']))

def mondays(start, end):
    d, out = start, []
    while d <= end:
        out.append(d); d += timedelta(days=7)
    return out

def month_firsts(start, n):
    out, y, m = [], start.year, start.month
    for _ in range(n):
        out.append(date(y, m, 1)); m += 1
        if m > 12: m, y = 1, y + 1
    return out

START = date(2026, 6, 15)
MAX_LEADS = 8
HITO_ROWS = 6          # filas de hitos (metas binarias con fecha) por pestaña de WIG
PCT = '0.0%'; CNT = '#,##0'; USD = '"$"#,##0'; DEC1 = '0.0'  # DEC1: conteos con 1 decimal (metas semanales convertidas de mensuales)

WIGS = [
    dict(tab='1. Smart User 500', title='WIG 1 — Incrementar el Parque de Smart User en 500 equipos en el 2026',
         lag='Equipos Smart User colocados (acumulado). 500 equipos a $40/mes = $6.4K MRR adicional en GP.',
         fromto='De 0 a 500 equipos colocados, antes del 31 de diciembre de 2026.',
         equipo='Pre-Sales, Del - DEX', meta=500, fmt=CNT, tipo='acum', end=date(2026, 12, 28), freq='W',
         # captura del lag: equipo = E4 del WIG; responsable que digita (fila 6) y fuente/método (fila 7)
         capta_resp='Sofía', fuente="Reporte 'on-hold' (semanal)",
         # captura por lead: (Equipo, Responsable, Método) — tres campos separados
         lead_capta=[('DEX Presale', 'Martha', 'Manual'),
                     ('AM / Contact Center', 'Ingrid', 'SSC'),
                     ('CS / AM', 'Margel', 'Manual'),
                     ('AM / Contact Center', 'Victor', 'Manual'),
                     ('Finanzas / Presale DEX', 'Federico', 'Manual'),
                     ('AM / DEX', 'Arlen', 'Manual'),
                     ('Carmen / Carolina', 'Contact Center', 'Presale DEX'),
                     ('AM / CS', 'Yenifer', 'Manual')],
         # hitos: metas binarias con fecha objetivo (logro = sí/no). (texto, lead, responsable, fecha, estado)
         hitos=[('Certificarse como Apple Authorized Service Provider', None, 'Del - DEX', date(2026, 9, 30), 'En curso'),
                ('Habilitar 50 equipos demo en showroom', 5, 'Finanzas / Presale DEX', date(2026, 7, 31), 'Pendiente')],
         # Leads reales (medidas predictivas). Meta = objetivo MENSUAL ÷ 4.33 → semanal,
         # excepto el nivel de inventario (50, se mantiene) y la tasa de acompañamiento (100%).
         leads=[('Presentación TCO', 0.7, DEC1),                         # 3/mes
                ('Llamadas en frío', 3.5, DEC1),                         # 15/mes
                ('Analizar parque restante', 0.2, DEC1),                 # 1/mes
                ('Entrega de ofertas no solicitadas', 3.5, DEC1),        # 15/mes
                ('Mantener inventario (nivel)', 50, CNT),                # nivel: 50 en bodega
                ('Acompañar cada oferta de venta directa con Smart User', 1.0, PCT),  # tasa: 100%
                ('Evento de generación de demanda', 0.2, DEC1),          # 1/mes
                ('Referencia de posibles clientes', 0.2, DEC1)]),        # 1/mes
    dict(tab='2. MV Advisory', title='WIG 2 — Lograr un Monthly Value de $X en advisory en el primer trimestre de 2027',
         lag='MV de horas advisory de ingenieros locales vendidas (USD acumulado). Costo fijo ya cubierto = mayor utilidad.',
         fromto='De $0 a la meta de MV advisory, antes del 31 de marzo de 2027. (Definir $X en la celda Meta)',
         equipo='Ventas, Customer Success, Pre-Sales, Del - TSS', meta=10000, fmt=USD, tipo='acum',
         end=date(2027, 3, 29), freq='W', meta_x=True,
         leads=[('Propuestas de advisory presentadas', 3, CNT),
                ('Horas advisory cotizadas', 40, CNT),
                ('Reuniones de venta consultiva con clientes', 4, CNT),
                ('Assessments / health checks ofrecidos', 2, CNT),
                ('% utilización de ingenieros locales en advisory', 0.20, PCT),
                ('Opps de advisory activas en pipeline', 6, CNT),
                ('Contratos advisory firmados en la semana', 1, CNT)]),
    dict(tab='3. Negocios +30K GP', title='WIG 3 — Reconocer 10 negocios mayores a $30K de GP en el 2027',
         lag='Negocios con GP > $30K reconocidos (acumulado). Negocios grandes que mueven el NAT.',
         fromto='De 0 a 10 negocios > $30K GP reconocidos, antes del 31 de diciembre de 2027.',
         equipo='Ventas, Pre-Sales', meta=10, fmt=CNT, tipo='acum', end=date(2027, 12, 27), freq='W',
         leads=[('Opps > $30K GP activas en pipeline', 5, CNT),
                ('Opps nuevas ingresadas en la semana', 4, CNT),
                ('Propuestas > $30K presentadas', 1, CNT),
                ('Reuniones C-level realizadas', 2, CNT),
                ('Leads activas al cierre de la semana', 10, CNT),
                ('Opps > $30K avanzadas de etapa', 2, CNT),
                ('Win rate de opps > $30K (acumulado)', 0.30, PCT)]),
    dict(tab='4. Entregas 100%', title='WIG 4 — Cumplir el 100% de entregas y colocaciones en tiempo y forma',
         lag='% de colocaciones y entregas de la semana realizadas en tiempo y forma.',
         fromto='Sostener 100% de cumplimiento semanal durante todo el periodo.',
         equipo='Finanzas', meta=1.0, fmt=PCT, tipo='nivel_min', end=date(2027, 12, 27), freq='W', riesgo=0.05,
         leads=[('% entregas programadas con checklist completo', 1.0, PCT),
                ('Riesgos de entrega escalados ≥1 semana antes', 1, CNT),
                ('% órdenes de compra liberadas a tiempo', 1.0, PCT),
                ('% inventario confirmado ≥5 días antes', 1.0, PCT),
                ('Coordinaciones logísticas confirmadas', 5, CNT),
                ('% facturas emitidas ≤48h tras la entrega', 1.0, PCT)]),
    dict(tab='5. MV Datacenter', title='WIG 5 — Incrementar el MV de servicios basados en el DC GBM San Marcos en $20,000',
         lag='MV nuevo firmado de servicios basados desde el datacenter (USD acumulado).',
         fromto='De $0 a +$20,000 de MV incremental, antes del 31 de diciembre de 2027.',
         equipo='Pre-Sales, Del - DC', meta=20000, fmt=USD, tipo='acum', end=date(2027, 12, 27), freq='W',
         leads=[('Propuestas de servicios DC presentadas', 2, CNT),
                ('Workloads nuevos aprovisionados / migrados', 1, CNT),
                ('Assessments de migración realizados', 1, CNT),
                ('Opps DC activas en pipeline', 4, CNT),
                ('Clientes contactados con oferta DC', 3, CNT),
                ('% capacidad DC documentada y actualizada', 1.0, PCT)]),
    dict(tab='6. Cero Expired', title='WIG 6 — Tener cero contratos expired en al menos 16 de los próximos 18 meses',
         lag='Contratos en estado Expired / On-Hold al cierre de la semana (meta: 0).',
         fromto='De N a 0 contratos expired, sosteniéndolo en 16 de los próximos 18 meses (jun 2026 – dic 2027).',
         equipo='Customer Success, Del - DEX, Del - TSS, Del - DC', meta=0, fmt=CNT, tipo='nivel_max',
         end=date(2027, 12, 27), freq='W', riesgo=1,
         leads=[('% contratos por vencer en 90 días con plan', 1.0, PCT),
                ('Renovaciones cerradas a tiempo en la semana', 2, CNT),
                ('Cotizaciones de renovación enviadas ≥60 días antes', 3, CNT),
                ('QBRs con clientes en riesgo de no renovar', 1, CNT),
                ('% contratos on-hold con plan de reactivación', 1.0, PCT),
                ('Forecast de renovaciones actualizado (1 = sí)', 1, CNT)]),
    dict(tab='7. MV Digital Sol', title='WIG 7 — Incrementar el MV de Digital Solutions en $30K/mes al 1ro de mayo 2027',
         lag='MV incremental firmado de servicios Digital Solutions (SOC y Célula), USD acumulado.',
         fromto='De $0 a +$30,000 de MV mensual, antes del 1 de mayo de 2027.',
         equipo='Ventas, Del - DS', meta=30000, fmt=USD, tipo='acum', end=date(2027, 4, 26), freq='W',
         leads=[('Opps SOC / Célula activas en pipeline', 4, CNT),
                ('Propuestas DS presentadas', 2, CNT),
                ('Demos / PoV de SOC ejecutados', 1, CNT),
                ('Workshops de células de desarrollo', 1, CNT),
                ('Casos de éxito / referencias compartidas', 1, CNT),
                ('Opps DS avanzadas de etapa', 2, CNT)]),
    dict(tab='8. GP 16%', title='WIG 8 — Lograr un GP consolidado de 16% en el 2027',
         lag='GP consolidado del mes (%).',
         fromto='De GP actual a 16% consolidado sostenido en 2027.',
         equipo='Finanzas (todas las áreas aportan)', meta=0.16, fmt=PCT, tipo='nivel_min',
         end=None, freq='M', riesgo=0.01,
         leads=[('% distribuciones del mes vs plan de baja distribución', 1.0, PCT),
                ('% reservas (inventario + ITA + IR) bajo el tope del plan', 1.0, PCT),
                ('% negocios cotizados con GP ≥ margen mínimo', 1.0, PCT),
                ('Acciones de recuperación de margen ejecutadas', 2, CNT),
                ('Revisiones de pricing realizadas en el mes', 4, CNT),
                ('% proyectos con revisión de rentabilidad mensual', 1.0, PCT)]),
    dict(tab='9. CSAT', title='WIG 9 — Incrementar el CSAT de X a Y, sosteniéndolo 4 meses consecutivos',
         lag='CSAT del mes (puntaje de encuestas de satisfacción).',
         fromto='De X a Y, con 4 meses consecutivos en la meta. (Definir X y Y en la celda Meta)',
         equipo='Customer Success (todas las áreas de entrega aportan)', meta=4.5, fmt='0.0', tipo='nivel_min',
         end=None, freq='M', riesgo=0.3, meta_x=True,
         leads=[('% encuestas CSAT enviadas y respondidas', 0.60, PCT),
                ('QBRs / touchpoints de CS ejecutados', 8, CNT),
                ('% detractores con plan de acción en ≤5 días', 1.0, PCT),
                ('% casos críticos resueltos dentro del SLA', 0.95, PCT),
                ('Llamadas proactivas de CSM en el mes', 12, CNT),
                ('Promotores contactados para referencia', 3, CNT)]),
    dict(tab='10. Apalancamiento', title='WIG 10 — Sostener costos locales planos mientras crece el ingreso (apalancamiento operativo)',
         lag='Costo indirecto local + OREX del mes como % del ingreso del mes (objetivo: que BAJE al crecer ventas).',
         fromto='De ~13% (local+OREX / ingreso) hacia ≤9% sostenido, manteniendo el costo local en valor absoluto plano mientras el ingreso crece.',
         equipo='Finanzas, Customer Success, todas las áreas locales', meta=0.09, fmt=PCT, tipo='nivel_max',
         end=None, freq='M', riesgo=0.01,
         leads=[('Costo local absoluto del mes vs. mes base (≤100%)', 1.0, PCT),
                ('Headcount local de apoyo vs. plan plano (≤100%)', 1.0, PCT),
                ('Vacantes de apoyo cubiertas con productividad, no contratación', 1, CNT),
                ('% de crecimiento de ingreso absorbido sin nuevo costo local', 1.0, PCT),
                ('Iniciativas de automatización/eficiencia en marcha', 1, CNT),
                ('Revisión mensual de costo local por área ejecutada (1 = sí)', 1, CNT)]),
    dict(tab='11. Recuperar DS', title='WIG 11 — Recuperar el margen de Digital Solutions y eliminar líneas con GP negativo',
         lag='GP del mes de las líneas hoy negativas (Digital Solutions + service-attach): de negativo a positivo (USD).',
         fromto='De ~-$78K/mes de arrastre (DS + líneas negativas) a GP ≥ $0 y luego positivo, antes de dic 2027.',
         equipo='Ventas, Pre-Sales, Del - DS, Finanzas', meta=0, fmt=USD, tipo='nivel_min',
         end=None, freq='M', riesgo=10000,
         leads=[('% negocios DS cotizados con GP ≥ piso de margen', 1.0, PCT),
                ('Líneas con GP negativo re-precificadas o cerradas (acum.)', 1, CNT),
                ('Contratos DS renegociados al alza en el mes', 1, CNT),
                ('% utilización de la célula / SOC', 0.70, PCT),
                ('Revisión de rentabilidad por línea DS ejecutada (1 = sí)', 1, CNT),
                ('Casos de service-attach negativo investigados', 2, CNT)]),
    dict(tab='12. Utilizacion', title='WIG 12 — Subir la utilización facturable de ingenieros locales de 42% a 48%+',
         lag='% de horas facturables sobre horas disponibles de los ingenieros locales (mes). Costo ya fijo = margen directo.',
         fromto='De 42% a ≥48% de utilización facturable, sostenido, antes de dic 2027.',
         equipo='Del - TSS, Del - DC, Customer Success, Pre-Sales', meta=0.48, fmt=PCT, tipo='nivel_min',
         end=None, freq='M', riesgo=0.03,
         leads=[('Horas facturables registradas en el mes vs. meta', 1.0, PCT),
                ('% de timesheets completos y a tiempo', 1.0, PCT),
                ('Horas advisory/billable vendidas en el mes', 40, CNT),
                ('Ingenieros con utilización < 40% con plan de acción', 1, CNT),
                ('Proyectos con staffing confirmado 2 semanas antes', 3, CNT),
                ('% de horas no facturables justificadas/categorizadas', 1.0, PCT)]),
]

wb = Workbook()
dash = wb.active
dash.title = 'Dashboard'
DATA0 = 11

def build_wig_tab(w):
    ws = wb.create_sheet(w['tab'])
    ws.sheet_view.showGridLines = False
    periods = mondays(START, w['end']) if w['freq'] == 'W' else month_firsts(date(2026, 7, 1), 18)
    plabel = 'Semana' if w['freq'] == 'W' else 'Mes'
    last = DATA0 + len(periods) - 1
    last_lead_col = 8 + 2 * MAX_LEADS

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_lead_col)
    ws['A1'] = w['title']; ws['A1'].font = F_TITLE
    def lbl(c, t): ws[c] = t; ws[c].font = F_LBL
    def txt(c, t, f=F_TXT): ws[c] = t; ws[c].font = f
    lbl('A2', 'Indicador lag:'); txt('B2', w['lag'])
    lbl('A3', 'De X a Y:'); txt('B3', w['fromto'])
    lbl('A4', 'Dueño:'); txt('B4', 'Por asignar', F_INPUT); ws['B4'].fill = FILL_YEL
    lbl('D4', 'Equipo:'); txt('E4', w['equipo'])
    lbl('A5', 'Meta:'); ws['B5'] = w['meta']; ws['B5'].font = F_INPUT; ws['B5'].number_format = w['fmt']
    if w.get('meta_x'): ws['B5'].fill = FILL_YEL
    lbl('D5', 'Inicio:'); ws['E5'] = periods[0]; ws['E5'].font = F_INPUT; ws['E5'].number_format = 'dd-mmm-yy'
    lbl('F5', 'Fin:'); ws['G5'] = periods[-1]; ws['G5'].font = F_INPUT; ws['G5'].number_format = 'dd-mmm-yy'
    # Captura del dato — tres campos: EQUIPO · RESPONSABLE (quién digita) · FUENTE/MÉTODO.
    # Lag (izquierda): equipo = E4 del WIG · responsable (B6) · fuente (B7).
    # Por lead (cols I+): fila 5 equipo · fila 6 responsable · fila 7 fuente.
    lbl('A6', 'Responsable:'); ws.merge_cells('B6:G6')
    cr = ws['B6']; cr.font = F_INPUT; cr.alignment = WRAP
    if w.get('capta_resp'): cr.value = w['capta_resp']
    lbl('A7', 'Fuente / método:'); ws.merge_cells('B7:G7')
    cfu = ws['B7']; cfu.font = F_INPUT; cfu.alignment = WRAP
    if w.get('fuente'): cfu.value = w['fuente']
    for rr, txt_h in ((5, 'Equipo →'), (6, 'Responsable →'), (7, 'Fuente →')):
        lc = ws.cell(row=rr, column=8, value=txt_h); lc.font = F_LBL
        lc.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[5].height = 24; ws.row_dimensions[6].height = 24; ws.row_dimensions[7].height = 24
    lbl('A8', 'Periodos:'); ws['B8'] = f'=COUNT(A{DATA0}:A{last})'; ws['B8'].font = F_FORM
    txt('A9', 'Celdas azules = se digitan cada lunes. Celdas negras = fórmulas, no tocar.', Font(name=ARIAL, size=9, italic=True, color='808080'))

    acum = w['tipo'] == 'acum'
    hdrs = ['#', plabel, ('Meta ' + plabel.lower()) if acum else 'Meta', 'Real',
            'Meta acum.' if acum else '', 'Real acum.' if acum else '',
            '% avance' if acum else 'Brecha vs meta', 'Estado']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=10, column=c, value=h)
        cell.font = F_HDR; cell.fill = FILL_HDR; cell.alignment = CENTER; cell.border = BORDER

    leads = list(w['leads'])[:MAX_LEADS]
    while len(leads) < MAX_LEADS: leads.append(('(Disponible)', None, CNT))
    lbln = ws.cell(row=8, column=8, value='Lead →'); lbln.font = F_LBL; lbln.alignment = Alignment(horizontal='right', vertical='center')
    lblm = ws.cell(row=9, column=8, value='Meta →'); lblm.font = F_LBL; lblm.alignment = Alignment(horizontal='right', vertical='center')
    for k, (name, meta, fmt) in enumerate(leads):
        c1 = 9 + 2 * k; c2 = c1 + 1
        ws.merge_cells(start_row=8, start_column=c1, end_row=8, end_column=c2)
        nc = ws.cell(row=8, column=c1, value=(f'L{k+1}. {name}' if name != '(Disponible)' else name))
        nc.font = F_INPUT_S; nc.alignment = WRAPC; nc.fill = FILL_GREY
        ws.merge_cells(start_row=9, start_column=c1, end_row=9, end_column=c2)
        mc = ws.cell(row=9, column=c1)
        if meta is not None: mc.value = meta
        mc.font = F_INPUT; mc.number_format = fmt; mc.alignment = CENTER
        # captura por lead: fila 5 = equipo, fila 6 = responsable, fila 7 = fuente / método
        lc = w.get('lead_capta') or []
        cap = lc[k] if k < len(lc) else None
        for rr, idx in ((5, 0), (6, 1), (7, 2)):
            ws.merge_cells(start_row=rr, start_column=c1, end_row=rr, end_column=c2)
            cc0 = ws.cell(row=rr, column=c1)
            if name != '(Disponible)' and cap and idx < len(cap) and cap[idx] is not None:
                cc0.value = cap[idx]
            cc0.font = F_INPUT_S; cc0.alignment = WRAPC; cc0.fill = FILL_GREY
        for rr in (5, 6, 7, 8, 9):
            for cc in (c1, c2): ws.cell(row=rr, column=cc).border = BORDER
        for cc, h in ((c1, 'Real'), (c2, '%')):
            cell = ws.cell(row=10, column=cc, value=h)
            cell.font = F_HDR_S; cell.fill = FILL_LEAD; cell.alignment = CENTER; cell.border = BORDER
    ws.row_dimensions[8].height = 42

    for i, p in enumerate(periods):
        r = DATA0 + i
        ws.cell(row=r, column=1, value=i + 1).font = F_FORM
        dc = ws.cell(row=r, column=2, value=p); dc.font = F_FORM; dc.number_format = 'dd-mmm-yy' if w['freq'] == 'W' else 'mmm-yy'
        if acum:
            ws.cell(row=r, column=3, value='=$B$5/$B$8').number_format = w['fmt']
            ws.cell(row=r, column=5, value=f'=SUM($C${DATA0}:C{r})').number_format = w['fmt']
            ws.cell(row=r, column=6, value=f'=SUM($D${DATA0}:D{r})').number_format = w['fmt']
            ws.cell(row=r, column=7, value=f'=IF(COUNT($D${DATA0}:D{r})=0,"",F{r}/E{r})').number_format = '0.0%'
            ws.cell(row=r, column=8, value=f'=IF(G{r}="","",IF(G{r}>=0.95,"En meta",IF(G{r}>=0.8,"Riesgo","Atrasado")))')
        else:
            ws.cell(row=r, column=3, value='=$B$5').number_format = w['fmt']
            ws.cell(row=r, column=7, value=f'=IF(D{r}="","",D{r}-C{r})').number_format = w['fmt']
            if w['tipo'] == 'nivel_min':
                ws.cell(row=r, column=8, value=f'=IF(D{r}="","",IF(D{r}>=C{r},"En meta",IF(D{r}>=C{r}-{w["riesgo"]},"Riesgo","Atrasado")))')
            else:
                ws.cell(row=r, column=8, value=f'=IF(D{r}="","",IF(D{r}<=C{r},"En meta",IF(D{r}<=C{r}+{w["riesgo"]},"Riesgo","Atrasado")))')
        for k in range(MAX_LEADS):
            c1 = 9 + 2 * k; L1 = col(c1); ML = f'${L1}$9'
            ws.cell(row=r, column=c1).number_format = leads[k][2]
            ws.cell(row=r, column=c1 + 1, value=f'=IF({L1}{r}="","",IF({ML}="","",IF({ML}=0,"",{L1}{r}/{ML})))').number_format = '0%'
        for c in range(1, last_lead_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            is_input = (c == 4) or (c >= 9 and (c - 9) % 2 == 0)
            if is_input: cell.font = F_INPUT_S if c >= 9 else F_INPUT
            else: cell.font = Font(name=ARIAL, size=9) if c >= 9 else F_FORM
            if c in (7, 8) or c >= 9: cell.alignment = CENTER
            if i % 2 == 1 and not is_input: cell.fill = FILL_GREY

    estado_cf(ws, f'H{DATA0}:H{last}')
    ws.freeze_panes = f'C{DATA0}'
    for cl, wd in {'A': 5, 'B': 11, 'C': 12, 'D': 11, 'E': 12, 'F': 12, 'G': 13, 'H': 11}.items():
        ws.column_dimensions[cl].width = wd
    for k in range(MAX_LEADS):
        ws.column_dimensions[col(9 + 2 * k)].width = 11
        ws.column_dimensions[col(10 + 2 * k)].width = 7

    ch = LineChart(); ch.title = 'Marcador: Meta vs Real'; ch.style = 12
    ch.height, ch.width = 8, 16
    if acum: data = Reference(ws, min_col=5, max_col=6, min_row=10, max_row=last)
    else: data = Reference(ws, min_col=3, max_col=4, min_row=10, max_row=last)
    cats = Reference(ws, min_col=2, min_row=DATA0, max_row=last)
    ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
    ch.series[0].graphicalProperties.line.solidFill = 'A6A6A6'
    ch.series[0].graphicalProperties.line.dashStyle = 'dash'
    ch.series[1].graphicalProperties.line.solidFill = '2E75B6'
    ch.series[1].graphicalProperties.line.width = 28000
    ws.add_chart(ch, f'{col(last_lead_col + 2)}2')

    # ---- Hitos: metas binarias con fecha objetivo (logro = sí/no) ----
    # Bloque debajo de la tabla de datos. El parser lo localiza por el rótulo
    # 'Hitos' en la col A tras el fin de los datos. Cols: A–D Hito (combinadas),
    # E Lead (1–8, opcional), F Responsable, G Fecha objetivo, H Estado
    # (Pendiente / En curso / Logrado). Lo digitan los dueños.
    hito0 = last + 2
    ws.merge_cells(start_row=hito0, start_column=1, end_row=hito0, end_column=last_lead_col)
    ws.cell(row=hito0, column=1,
            value='Hitos — metas binarias con fecha objetivo (logro = sí/no)').font = F_SECT
    hh = hito0 + 1
    ws.merge_cells(start_row=hh, start_column=1, end_row=hh, end_column=4)
    for c, h in ((1, 'Hito'), (5, 'Lead'), (6, 'Responsable'), (7, 'Fecha objetivo'), (8, 'Estado')):
        cell = ws.cell(row=hh, column=c, value=h)
        cell.font = F_HDR; cell.fill = FILL_HDR; cell.alignment = CENTER; cell.border = BORDER
    hr0, hr1 = hh + 1, hh + HITO_ROWS
    seed = list(w.get('hitos') or [])
    dv_hl = DataValidation(type='list', formula1='"%s"' % ','.join(str(i) for i in range(1, MAX_LEADS + 1)), allow_blank=True)
    dv_he = DataValidation(type='list', formula1='"Pendiente,En curso,Logrado"', allow_blank=True)
    ws.add_data_validation(dv_hl); ws.add_data_validation(dv_he)
    dv_hl.add(f'E{hr0}:E{hr1}'); dv_he.add(f'H{hr0}:H{hr1}')
    for j in range(HITO_ROWS):
        r = hr0 + j
        s = seed[j] if j < len(seed) else (None,) * 5
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        vals = {1: s[0], 5: s[1], 6: s[2], 7: s[3], 8: s[4]}
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            if c in vals: cell.value = vals[c]
            cell.border = BORDER; cell.font = F_INPUT
            cell.alignment = WRAP if c == 1 else CENTER
            if c == 7: cell.number_format = 'dd-mmm-yy'
        ws.row_dimensions[r].height = 22
    for txt_, dxf in (('Logrado', GREEN_DXF), ('En curso', YEL_DXF), ('Pendiente', RED_DXF)):
        ws.conditional_formatting.add(f'H{hr0}:H{hr1}', Rule(type='containsText', operator='containsText',
            text=txt_, dxf=dxf, formula=[f'NOT(ISERROR(SEARCH("{txt_}",H{hr0})))']))
    return last, hr0, hr1

built = {w['tab']: build_wig_tab(w) for w in WIGS}
lasts = {t: v[0] for t, v in built.items()}
hitos_ranges = {t: (v[1], v[2]) for t, v in built.items()}

# ---------- Páginas por año (meta de utilidad neta + cobertura de backlog) ----------
# Una pestaña por año futuro (2027/2028/2029). Encabeza con la META DE UTILIDAD
# NETA (NAT) del año; debajo, el camino de GP y la cobertura de backlog: cuánto
# GP ya está comprometido por contratos conocidos y cuánto falta vender (brecha).
# La tabla semanal se actualiza cada lunes de 2026.
BACKLOG_START = date(2026, 6, 15)
BACKLOG_END = date(2026, 12, 28)
BL_DATA0 = 12
BACKLOG_YEARS = [
    dict(tab='2027', year='2027', ingreso=21000000, gp_pct=0.16, nat_meta=1000000, assume_gp=False),
    dict(tab='2028', year='2028', ingreso=28571000, gp_pct=0.164, nat_meta=2000000, assume_gp=False),
    # 2029: vende $2M más que 2028; margen neto 8%; GP% es supuesto editable.
    dict(tab='2029', year='2029', ingreso=30571000, gp_pct=0.165, nat_meta=2445680, assume_gp=True),
]

def build_year_tab(b):
    ws = wb.create_sheet(b['tab'])
    ws.sheet_view.showGridLines = False
    periods = mondays(BACKLOG_START, BACKLOG_END)
    last = BL_DATA0 + len(periods) - 1
    nat_pct = b['nat_meta'] / b['ingreso']
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{b['year']} — Meta de Utilidad Neta (NAT)"; ws['A1'].font = F_TITLE
    ws.merge_cells('A2:F2')
    ws['A2'] = (f"Meta de utilidad neta después de impuestos para {b['year']}. Debajo: el camino de GP y la "
                f"cobertura de backlog — cuánto GP ya está comprometido por contratos conocidos y cuánto "
                f"falta vender. Se actualiza cada semana de {BACKLOG_START.year}.")
    ws['A2'].font = F_TXT; ws['A2'].alignment = WRAP
    def lbl(c, t): ws[c] = t; ws[c].font = F_LBL
    def numc(c, v, f=USD, font=F_FORM):
        cell = ws[c]; cell.value = v; cell.font = font; cell.number_format = f
    lbl('A4', 'Ingreso meta:'); numc('B4', b['ingreso'])
    lbl('D4', 'GP meta %:')
    gp = ws['E4']; gp.value = b['gp_pct']; gp.number_format = PCT
    gp.font = F_INPUT if b['assume_gp'] else F_FORM
    if b['assume_gp']: gp.fill = FILL_YEL  # supuesto a confirmar
    lbl('A5', 'GP meta (Ingreso × GP%):'); numc('B5', '=B4*E4')
    lbl('D5', 'META NAT (utilidad neta):')
    natc = ws['E5']; natc.value = b['nat_meta']; natc.number_format = USD
    natc.font = Font(name=ARIAL, size=14, bold=True, color='1F3864')  # hero: la meta del año
    lbl('A6', 'GP comprometido (último):')
    numc('B6', f'=IFERROR(LOOKUP(2,1/($B${BL_DATA0}:$B${last}<>""),$B${BL_DATA0}:$B${last}),0)')
    lbl('D6', '% cobertura:'); ws['E6'].value = '=IFERROR(B6/B5,0)'; ws['E6'].number_format = PCT; ws['E6'].font = F_FORM
    lbl('A7', 'Brecha (falta vender):'); numc('B7', '=B5-B6')
    lbl('D7', 'Margen NAT (NAT / Ingreso):'); ws['E7'].value = nat_pct; ws['E7'].number_format = PCT; ws['E7'].font = F_FORM
    ws['A9'] = f"Backlog comprometido por semana ({b['year']})"
    ws['A9'].font = Font(name=ARIAL, size=12, bold=True, color='1F3864')
    hdrs = ['Semana', 'GP comprometido (acum.)', 'GP meta', 'Brecha', '% cobertura']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=11, column=c, value=h)
        cell.font = F_HDR; cell.fill = FILL_HDR; cell.alignment = WRAPC; cell.border = BORDER
    for i, p in enumerate(periods):
        r = BL_DATA0 + i
        dc = ws.cell(row=r, column=1, value=p); dc.font = F_FORM; dc.number_format = 'dd-mmm-yy'
        bc = ws.cell(row=r, column=2); bc.font = F_INPUT; bc.number_format = USD  # GP comprometido (acum., entrada)
        ws.cell(row=r, column=3, value='=$B$5').number_format = USD
        ws.cell(row=r, column=4, value=f'=IF(B{r}="","",C{r}-B{r})').number_format = USD
        ws.cell(row=r, column=5, value=f'=IF(B{r}="","",B{r}/C{r})').number_format = '0.0%'
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c); cell.border = BORDER
            if c in (3, 4, 5): cell.alignment = CENTER
            if i % 2 == 1 and c != 2: cell.fill = FILL_GREY
    ws.freeze_panes = f'A{BL_DATA0}'
    for cl, wd in {'A': 12, 'B': 22, 'C': 14, 'D': 14, 'E': 12}.items():
        ws.column_dimensions[cl].width = wd
    ch = LineChart(); ch.title = f'Cobertura {b["year"]}: GP comprometido vs meta'; ch.style = 12
    ch.height, ch.width = 8, 16
    data = Reference(ws, min_col=2, max_col=3, min_row=11, max_row=last)
    cats = Reference(ws, min_col=1, min_row=BL_DATA0, max_row=last)
    ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
    ch.series[0].graphicalProperties.line.solidFill = '2E75B6'
    ch.series[0].graphicalProperties.line.width = 28000
    ch.series[1].graphicalProperties.line.solidFill = 'A6A6A6'
    ch.series[1].graphicalProperties.line.dashStyle = 'dash'
    ws.add_chart(ch, 'G4')
    return last

year_lasts = {b['tab']: build_year_tab(b) for b in BACKLOG_YEARS}

# ---------- Tareas de soporte (por lead) ----------
# Tabla plana editable: cada fila es una tarea/iniciativa concreta que sostiene
# un lead measure específico. WIG = número de pestaña (1–12), Lead = 1–8. El
# marcador la lee por posición A–E y la muestra en el detalle de cada lead.
# Los dueños agregan filas en la copia compartida; los desplegables evitan
# números fuera de rango. La semilla documenta el formato con ejemplos reales.
# Nota: las metas binarias con fecha (p. ej. "Certificarse como Apple Authorized
# Service Provider") ya no van aquí; viven en el bloque "Hitos" de cada pestaña de WIG.
TAREAS_SEED = [
    (1, 3, 'Presentar la oferta Smart User a 100 clientes en 2026', 'Pre-Sales', 'En curso'),
    (1, 5, 'Mantener siempre 50 equipos en bodega listos para colocar', 'Del - DEX', 'En curso'),
]
TAREAS_ROWS = 120  # filas de entrada disponibles para los dueños

def build_tareas_tab():
    ws = wb.create_sheet('Tareas')
    ws.sheet_view.showGridLines = False
    ws['G1'] = ('Tareas de soporte por lead: cada fila es una acción concreta que sostiene un '
                'lead measure. WIG = número de pestaña (1–12), Lead = 1–8. Aparecen en el detalle '
                'del lead en el televisor. Estado: Pendiente · En curso · Hecho.')
    ws['G1'].font = Font(name=ARIAL, size=9, italic=True, color='808080'); ws['G1'].alignment = WRAP
    hdrs = ['WIG', 'Lead', 'Tarea', 'Responsable', 'Estado']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = F_HDR; cell.fill = FILL_HDR; cell.alignment = CENTER; cell.border = BORDER
    last = 1 + TAREAS_ROWS
    # desplegables (validación) para WIG, Lead y Estado
    dv_wig = DataValidation(type='list', formula1='"%s"' % ','.join(str(i) for i in range(1, len(WIGS) + 1)), allow_blank=True)
    dv_lead = DataValidation(type='list', formula1='"%s"' % ','.join(str(i) for i in range(1, MAX_LEADS + 1)), allow_blank=True)
    dv_est = DataValidation(type='list', formula1='"Pendiente,En curso,Hecho"', allow_blank=True)
    for dv in (dv_wig, dv_lead, dv_est): ws.add_data_validation(dv)
    dv_wig.add(f'A2:A{last}'); dv_lead.add(f'B2:B{last}'); dv_est.add(f'E2:E{last}')
    for i in range(TAREAS_ROWS):
        r = 2 + i
        seed = TAREAS_SEED[i] if i < len(TAREAS_SEED) else (None,) * 5
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c, value=seed[c - 1])
            cell.border = BORDER; cell.font = F_INPUT
            cell.alignment = CENTER if c in (1, 2, 5) else WRAP
            if i % 2 == 1 and cell.value is None: cell.fill = FILL_GREY
    # color de Estado: Hecho = verde · En curso = amarillo · Pendiente = rojo
    for txt, dxf in (('Hecho', GREEN_DXF), ('En curso', YEL_DXF), ('Pendiente', RED_DXF)):
        ws.conditional_formatting.add(f'E2:E{last}', Rule(type='containsText', operator='containsText',
            text=txt, dxf=dxf, formula=[f'NOT(ISERROR(SEARCH("{txt}",E2)))']))
    ws.freeze_panes = 'A2'
    for cl, wd in {'A': 6, 'B': 6, 'C': 72, 'D': 22, 'E': 12}.items():
        ws.column_dimensions[cl].width = wd
    return last

build_tareas_tab()

# ---------- Dashboard ----------
ws = dash
ws.sheet_view.showGridLines = False
ws.merge_cells('A1:H1'); ws['A1'] = 'Tablero 4DX — GBM Nicaragua'; ws['A1'].font = Font(name=ARIAL, size=16, bold=True, color='1F3864')
ws.merge_cells('A2:H2'); ws['A2'] = 'WIG de la compañía: Construir la utilidad neta (NAT) de Nicaragua — $1M en 2027, $2M (7%) en 2028, $2.4M (8%) en 2029'
ws['A2'].font = Font(name=ARIAL, size=11, bold=True)

# --- Metas de Utilidad Neta (NAT) 2027 → 2029, con base 2026 (NAT real = entrada anual) ---
ws['A4'] = 'Metas de Utilidad Neta (NAT) 2027 → 2029  (base 2026)'; ws['A4'].font = F_SECT
traj_hdrs = ['Año', 'Ingreso', 'GP %', 'Meta NAT', 'NAT %', 'NAT real', 'Estado', 'Ir a pestaña']
for c, h in enumerate(traj_hdrs, 1):
    cell = ws.cell(row=5, column=c, value=h)
    cell.font = F_HDR; cell.fill = FILL_HDR; cell.alignment = CENTER; cell.border = BORDER
# (etiqueta, ingreso, GP%, NAT meta, pestaña del año o None para la base)
TRAJ = [('2026 (estimado)', 18656000, 0.144, 375000, None),
        ('2027 (meta)', 21000000, 0.16, 1000000, '2027'),
        ('2028 (meta)', 28571000, 0.164, 2000000, '2028'),
        ('2029 (meta)', 30571000, 0.165, 2445680, '2029')]  # 2029: +$2M ventas vs 2028, NAT 8%
TRAJ0 = 6
for i, (yr, ing, gp, nat, tab) in enumerate(TRAJ):
    r = TRAJ0 + i
    ws.cell(row=r, column=1, value=yr).font = F_FORM
    ws.cell(row=r, column=2, value=ing).number_format = USD
    ws.cell(row=r, column=3, value=gp).number_format = PCT
    ws.cell(row=r, column=4, value=nat).number_format = USD
    ws.cell(row=r, column=5, value=f'=D{r}/B{r}').number_format = PCT
    rc = ws.cell(row=r, column=6); rc.font = F_INPUT; rc.number_format = USD  # NAT real (entrada)
    ws.cell(row=r, column=7, value=f'=IF(F{r}="","—",IF(F{r}>=D{r}*0.95,"En meta",IF(F{r}>=D{r}*0.8,"Riesgo","Atrasado")))')
    if tab:
        lk = ws.cell(row=r, column=8, value=f'Ver {tab}'); lk.hyperlink = f"#'{tab}'!A1"
        lk.font = Font(name=ARIAL, size=10, color='0563C1', underline='single')
    for c in range(1, 9):
        cell = ws.cell(row=r, column=c); cell.border = BORDER
        if c not in (6, 8) and cell.font.color is None: cell.font = F_FORM
        if c in (3, 5, 7, 8): cell.alignment = CENTER
traj_last = TRAJ0 + len(TRAJ) - 1
estado_cf(ws, f'G{TRAJ0}:G{traj_last}')
ws.cell(row=10, column=1, value='El tope de impuesto mínimo (3% sobre ventas) fija el piso de impuestos; por eso la meta se gana en GP% y en mantener el costo local plano, no solo en crecer ventas.').font = Font(name=ARIAL, size=9, italic=True, color='808080')

# --- Cobertura de backlog: GP ya comprometido vs meta del año (lee las páginas Backlog) ---
ws['A12'] = 'Cobertura de backlog — GP comprometido vs meta del año'; ws['A12'].font = F_SECT
bl_hdrs = ['Año', 'GP meta', 'GP comprometido', 'Brecha (falta vender)', '% cobertura']
for c, h in enumerate(bl_hdrs, 1):
    cell = ws.cell(row=13, column=c, value=h)
    cell.font = F_HDR; cell.fill = FILL_HDR; cell.alignment = WRAPC; cell.border = BORDER
for i, b in enumerate(BACKLOG_YEARS):
    r = 14 + i
    q = f"'{b['tab']}'"
    ws.cell(row=r, column=1, value=int(b['year'])).font = F_FORM
    ws.cell(row=r, column=2, value=f'={q}!$B$5').number_format = USD       # GP meta
    ws.cell(row=r, column=3, value=f'={q}!$B$6').number_format = USD       # comprometido último
    ws.cell(row=r, column=4, value=f'={q}!$B$7').number_format = USD       # brecha
    ws.cell(row=r, column=5, value=f'={q}!$E$6').number_format = PCT       # % cobertura
    for c in range(1, 6):
        cell = ws.cell(row=r, column=c); cell.border = BORDER
        if cell.font.color is None: cell.font = F_LINK if c >= 2 else F_FORM
        if c in (1, 5): cell.alignment = CENTER
        if i % 2 == 1: cell.fill = FILL_GREY
bl_last = 14 + len(BACKLOG_YEARS) - 1

# --- Seguimiento mensual NAT del año en curso ---
ws['A18'] = 'Seguimiento mensual NAT — año en curso'; ws['A18'].font = F_SECT
ws['A19'] = 'Meta anual NAT:'; ws['A19'].font = F_LBL
ws['C19'] = 1000000; ws['C19'].font = F_INPUT; ws['C19'].number_format = USD; ws['C19'].fill = FILL_YEL
DASH0 = 22
hdrs = ['Mes', 'Meta mensual', 'NAT real', 'Meta acum.', 'Real acum.', '% avance', 'Estado']
for c, h in enumerate(hdrs, 1):
    cell = ws.cell(row=21, column=c, value=h)
    cell.font = F_HDR; cell.fill = FILL_HDR; cell.alignment = CENTER; cell.border = BORDER
for i, m in enumerate(month_firsts(date(2027, 1, 1), 12)):
    r = DASH0 + i
    mc = ws.cell(row=r, column=1, value=m); mc.number_format = 'mmm-yy'; mc.font = F_FORM
    ws.cell(row=r, column=2, value='=$C$19/12').number_format = USD
    rc = ws.cell(row=r, column=3); rc.font = F_INPUT; rc.number_format = USD
    ws.cell(row=r, column=4, value=f'=SUM($B${DASH0}:B{r})').number_format = USD
    ws.cell(row=r, column=5, value=f'=SUM($C${DASH0}:C{r})').number_format = USD
    ws.cell(row=r, column=6, value=f'=IF(COUNT($C${DASH0}:C{r})=0,"",E{r}/D{r})').number_format = '0.0%'
    ws.cell(row=r, column=7, value=f'=IF(F{r}="","",IF(F{r}>=0.95,"En meta",IF(F{r}>=0.8,"Riesgo","Atrasado")))')
    for c in range(1, 8):
        cell = ws.cell(row=r, column=c); cell.border = BORDER
        if c != 3 and cell.font.color is None: cell.font = F_FORM
        if c in (6, 7): cell.alignment = CENTER
        if i % 2 == 1 and c != 3: cell.fill = FILL_GREY
last_dash = DASH0 + 11
estado_cf(ws, f'G{DASH0}:G{last_dash}')
ch = LineChart(); ch.title = 'NAT 2027 — Meta vs Real (acumulado)'; ch.style = 12
ch.height, ch.width = 8, 14
data = Reference(ws, min_col=4, max_col=5, min_row=21, max_row=last_dash)
cats = Reference(ws, min_col=1, min_row=DASH0, max_row=last_dash)
ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
ch.series[0].graphicalProperties.line.solidFill = 'A6A6A6'
ch.series[0].graphicalProperties.line.dashStyle = 'dash'
ch.series[1].graphicalProperties.line.solidFill = '2E75B6'
ch.series[1].graphicalProperties.line.width = 28000
ws.add_chart(ch, 'I18')
R0 = 36
ws.cell(row=R0 - 1, column=1, value='WIGs de soporte — estado al último dato ingresado').font = F_SECT
hdrs = ['#', 'WIG de soporte', 'Dueño', 'Equipo', 'Meta', 'Último real', 'Estado', 'Captura del lag', 'Ir a pestaña']
for c, h in enumerate(hdrs, 1):
    cell = ws.cell(row=R0, column=c, value=h)
    cell.font = F_HDR; cell.fill = FILL_HDR; cell.alignment = WRAPC; cell.border = BORDER
for i, w in enumerate(WIGS):
    r = R0 + 1 + i
    t, last = w['tab'], lasts[w['tab']]
    q = f"'{t}'"
    ws.cell(row=r, column=1, value=i + 1).font = F_FORM
    nm = ws.cell(row=r, column=2, value=w['title'].split('— ')[1]); nm.font = F_FORM; nm.alignment = WRAP
    ws.cell(row=r, column=3, value=f'={q}!$B$4').font = F_LINK
    eq = ws.cell(row=r, column=4, value=w['equipo']); eq.font = F_TXT; eq.alignment = WRAP
    mc = ws.cell(row=r, column=5, value=f'={q}!$B$5'); mc.font = F_LINK; mc.number_format = w['fmt']
    lr = ws.cell(row=r, column=6, value=f'=IFERROR(LOOKUP(2,1/({q}!$D${DATA0}:$D${last}<>""),{q}!$D${DATA0}:$D${last}),"Sin datos")')
    lr.font = F_LINK; lr.number_format = w['fmt']
    es = ws.cell(row=r, column=7, value=f'=IFERROR(LOOKUP(2,1/({q}!$D${DATA0}:$D${last}<>""),{q}!$H${DATA0}:$H${last}),"Sin datos")')
    es.font = F_LINK; es.alignment = CENTER
    # captura del lag: equipo (E4) · responsable (B6) — fuente (B7) de la pestaña del WIG.
    # Concatenación con & (evita TEXTJOIN, que LibreOffice headless no recalcula).
    cap = ws.cell(row=r, column=8,
                  value=f'={q}!$E$4&IF({q}!$B$6=""," "," · "&{q}!$B$6)&IF({q}!$B$7=""," "," — "&{q}!$B$7)')
    cap.font = F_LINK; cap.alignment = WRAP
    lk = ws.cell(row=r, column=9, value=t); lk.hyperlink = f"#'{t}'!A1"; lk.font = Font(name=ARIAL, size=10, color='0563C1', underline='single')
    for c in range(1, 10):
        ws.cell(row=r, column=c).border = BORDER
        if i % 2 == 1: ws.cell(row=r, column=c).fill = FILL_GREY
estado_cf(ws, f'G{R0 + 1}:G{R0 + len(WIGS)}')
for cl, wd in {'A': 5, 'B': 46, 'C': 16, 'D': 26, 'E': 11, 'F': 11, 'G': 11, 'H': 34, 'I': 16}.items():
    ws.column_dimensions[cl].width = wd
for i in range(len(WIGS)): ws.row_dimensions[R0 + 1 + i].height = 30

# --- Hitos por WIG — metas binarias con fecha (Logrado / En curso / Pendiente) ---
HR = R0 + len(WIGS) + 3   # deja una fila en blanco bajo la tabla de soporte
ws.cell(row=HR, column=1, value='Hitos por WIG — metas binarias con fecha objetivo').font = F_SECT
h_hdrs = ['#', 'WIG', 'Logrado', 'En curso', 'Pendiente', 'Próxima fecha']
for c, h in enumerate(h_hdrs, 1):
    cell = ws.cell(row=HR + 1, column=c, value=h)
    cell.font = F_HDR; cell.fill = FILL_HDR; cell.alignment = WRAPC; cell.border = BORDER
for i, w in enumerate(WIGS):
    r = HR + 2 + i
    t = w['tab']; q = f"'{t}'"
    g0, g1 = hitos_ranges[t]
    ws.cell(row=r, column=1, value=i + 1).font = F_FORM
    nm = ws.cell(row=r, column=2, value=w['title'].split('— ')[1]); nm.font = F_FORM; nm.alignment = WRAP
    for c, est in ((3, 'Logrado'), (4, 'En curso'), (5, 'Pendiente')):
        cc = ws.cell(row=r, column=c, value=f'=COUNTIF({q}!$H${g0}:$H${g1},"{est}")')
        cc.font = F_LINK; cc.alignment = CENTER
    fc = ws.cell(row=r, column=6, value=f'=IF(COUNT({q}!$G${g0}:$G${g1})=0,"",MIN({q}!$G${g0}:$G${g1}))')
    fc.font = F_LINK; fc.number_format = 'dd-mmm-yy'; fc.alignment = CENTER
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = BORDER
        if i % 2 == 1: ws.cell(row=r, column=c).fill = FILL_GREY
for i in range(len(WIGS)): ws.row_dimensions[HR + 2 + i].height = 26

# ---------- Instrucciones ----------
ins = wb.create_sheet('Instrucciones')
ins.sheet_view.showGridLines = False
ins.column_dimensions['A'].width = 110
lines = [
    ('Cómo usar este tablero 4DX', F_TITLE), ('', None),
    ('RITUAL DEL LUNES (antes de la reunión de WIG):', F_LBL),
    ('1. Cada dueño de WIG abre su pestaña y digita en las celdas AZULES de la semana: el resultado real del lag (columna "Real") y el real de cada lead measure (columnas "Real" bajo cada lead, a la derecha).', F_TXT),
    ('2. Finanzas digita el NAT del mes en el Dashboard (columna "NAT real") cuando cierre el mes.', F_TXT),
    ('3. El Dashboard se actualiza solo, y el marcador del televisor se refresca arrastrando este archivo a la pantalla.', F_TXT),
    ('4. Suban el archivo actualizado a Claude cada lunes y pidan el marcador semanal para la reunión de cadencia.', F_TXT),
    ('', None),
    ('LEAD MEASURES (hasta 8 por WIG):', F_LBL),
    ('• Cada WIG tiene hasta 8 indicadores predictivos en columnas a la derecha: el nombre (fila 8) y la meta semanal/mensual (fila 9) son editables en azul.', F_TXT),
    ('• Los leads L1 y L2 son la apuesta principal: ahí el equipo rinde cuentas con compromisos semanales. L3–L8 son indicadores de apoyo que explican el resultado.', F_TXT),
    ('• El % de cada lead se calcula automáticamente contra su meta. Dejar la meta vacía desactiva el lead ("(Disponible)").', F_TXT),
    ('• Definan todos los leads en positivo (más = mejor) para que el % se lea igual en todos.', F_TXT),
    ('', None),
    ('TAREAS DE SOPORTE (pestaña «Tareas»):', F_LBL),
    ('• Cada fila es una acción concreta que sostiene un lead measure: WIG (1–12), Lead (1–8), la tarea, el responsable y el estado.', F_TXT),
    ('• Usen los desplegables de las columnas WIG, Lead y Estado (Pendiente · En curso · Hecho) para no salirse de rango.', F_TXT),
    ('• Aparecen en el televisor al tocar el lead correspondiente (sección «Tareas de soporte» del detalle).', F_TXT),
    ('', None),
    ('CÓDIGO DE COLORES:', F_LBL),
    ('• Texto AZUL = celdas de entrada. Texto NEGRO = fórmulas (no tocar). Texto VERDE = vínculos entre pestañas.', F_TXT),
    ('• Fondo AMARILLO = supuestos pendientes de definir (dueños, metas "X" de los WIGs 2 y 9).', F_TXT),
    ('• Estado: En meta ≥ 95% del plan · Riesgo 80–95% · Atrasado < 80%.', F_TXT),
    ('', None),
    ('PRINCIPIO 4DX:', F_LBL),
    ('Las lead measures son predictivas e influenciables por el equipo; el lag solo se mueve si las leads se cumplen. En la cadencia semanal cada quien rinde cuentas de sus compromisos sobre las leads, no sobre el lag.', F_TXT),
]
for i, (t, f) in enumerate(lines, 1):
    if t:
        ins.cell(row=i, column=1, value=t).font = f
        ins.cell(row=i, column=1).alignment = WRAP

import sys
out = sys.argv[1] if len(sys.argv) > 1 else 'dist/Tablero_WIG_4DX_GBM.xlsx'
wb.save(out)
print('built', out)
